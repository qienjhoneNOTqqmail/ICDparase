

from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.styles import Color, Font, Side, Border, PatternFill, Style
from openpyxl.reader.excel import load_workbook
from bunch import Bunch

yellow   = Color(rgb='00ffff00')
thinline = Side(style='thin')
border   = Border(left=thinline, right=thinline, top=thinline, bottom=thinline)
hdrstyle = Style(
    font=Font(b=True),
    fill=PatternFill(patternType='solid', bgColor=yellow, fgColor=yellow),
    border=border
)

normstyle=Style(border=border)

def genExcelSheet(wb, title, descriptor, data):
    ws = wb.create_sheet(title=title)
    # write header 
    ws.append([col[0] for col in descriptor])
    # write data 
    for rec in data:
        ws.append([rec.get(col[1],"") for col in descriptor])

    # set column width
    colidx = 1
    for coldesc in descriptor:
        letter = get_column_letter(colidx)
        ws.column_dimensions[letter].width = coldesc[2]
        colidx += 1
    
    # set styles
    rowstyle = hdrstyle
    for row in ws.rows:
        for cell in row:
            cell.style = rowstyle
        rowstyle = normstyle
        
    # freeze line/column 1
    ws._freeze_panes = "B2"
        
def genExcelFile(filename, sheets):
    wb = Workbook()
    del wb.worksheets[0]
    for title, descriptor, data in sheets:
        genExcelSheet(wb, title, descriptor, data)
    wb.save(filename)
    
def getExcelFileSheetNames(filename):
    wb = load_workbook(filename, read_only=True, data_only=True)

    return wb.sheetnames

def readExcelFile(filename, sheets):
    wb = load_workbook(filename, read_only=True, data_only=True)

    resultlst = []
    for title in sheets:
        ws = wb.get_sheet_by_name(title)
        if not ws:
            return None

        rowidx = 0
        reclst = []
        for row in ws.iter_rows():
            if rowidx == 0:
                # read header
                header = [str(cell.value).strip() for cell in row]
            else:
                # get a line, skip if first column starts with #
                rec = Bunch()
                colidx = 0
                for cell in row:
                    rec[header[colidx]] = cell.value
                    colidx += 1
                
                skip = rec.get('Skip')
                if skip and skip.startswith('#'):
                    continue
                reclst.append(rec)
            
            rowidx += 1

        resultlst.append(reclst)
        
    return resultlst

if __name__ == "__main__":
    x = readExcelFile("HF_IDUCENTER-icd.xlsx", ('InputMessages', 'InputSignals'))
    print len(x)
    for lst in x:
        print len(lst)
    
            
        
        
        
                    
                

            

                
        
        
