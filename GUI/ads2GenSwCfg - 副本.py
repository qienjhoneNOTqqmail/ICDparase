from openpyxl import load_workbook, worksheet, Workbook
import re
import sys
import os.path
import iomGenCommonV2

SW_CFG_TMPL = '''
disable edp ports all
disable learning  ports all
disable flooding unicast ports all
disable flooding multicast  ports all
delete fdb all
delete vlan vafdx
'''

def getIdxPort(app, vlan_list):

	for i in range(len(vlan_list)):
		if app in vlan_list[i][2]:
			return i
	return -1

def createOutput(outfn, s):
   #os.makedirs(os.path.dirname(outfn))
   open(outfn, "w").write(s)
	
def genSwCfg(data, vlan_lst, outfn):

	s=SW_CFG_TMPL

	s+= "create vlan vafdx\n"
	s+= "configure vafdx tag 100\n"
	s+= "configure default delete port 1-48\n"
	s+= "configure vafdx add ports 1-48\n"
	s+= "enable learning  port 1-48\n\n"

	vlsPorts = {}
	
	#DS AFDX
	for app in data:
		print "configure sw for %s" % (app[0])
		for vl in app[1].afdx.vls.values(): #input vl
			print "-check in vl %s" % (vl.vlName)
			port =  vlan_lst[getIdxPort(app[0], vlan_lst)][0]
			if not port:
				continue
			if vl.macAdd not in vlsPorts.keys():
				vlsPorts[vl.macAdd] = []
			if not str(port) in vlsPorts[vl.macAdd]:
				vlsPorts[vl.macAdd].append(str(port))
		for vl in app[2].afdx.vls.values(): #output vl
			print "-check in vl %s" % (vl.vlName)
			port =  vlan_lst[getIdxPort(app[0], vlan_lst)][1]
			if not port:
				continue
			if vl.macAdd not in vlsPorts.keys():
				vlsPorts[vl.macAdd] = []
			if not str(port) in vlsPorts[vl.macAdd]:
				vlsPorts[vl.macAdd].append(str(port))
				
	for vl in vlsPorts.keys():
		print vl + " " + repr(vlsPorts[vl])
		p = ""
		s+="Create fdbentry "+vl+" vlan vafdx ports %s\n"%(( ",".join(vlsPorts[vl])))	
				
	createOutput(outfn, s)

def main(args):
	if len(args) <= 3:
		sys.stderr.write("Usage: ads2GenSwCfg outputdir deviceMapp excelIcd1 excelIcd2 ... ")
		sys.exit(1)
	
	outputdir = args[0]
	deviceMap = args[1]
	inputfiles = args[2:]
	
	data = []
	lrus = []
	
	#read excel icds
	for inputfile in inputfiles:
		class_ = os.path.basename(inputfile).split('.')[0]
		class_ = class_.replace("-icd", "")
		
		modelIn = iomGenCommonV2.MapReader(inputfile, msgonly=True, direction='input')
		modelOut = iomGenCommonV2.MapReader(inputfile, msgonly=True, direction='output')
		data.append([class_, modelIn, modelOut])
		lrus.append(class_)
	
	#assamble vlan list
	vlan_lst=[]
	devMapIdx={"LRU":-1,"SW_PORT_SIM":-1,"SW_PORT_UUT":-1}
	wb = load_workbook(deviceMap)
	devMap = wb.get_sheet_by_name(name = 'DEVICE_MAP')
	for row in devMap.rows:
		for cell in row:
			for k in devMapIdx.keys():
				if devMapIdx[k] != -1:
					continue
				else:
					if str(cell.value).startswith(k):
						devMapIdx[k] = ord(cell.column)-65
						break
		if -1 not in devMapIdx.values():
			break
	for mapping in devMap.iter_rows(row_offset=1):
		lru = str(mapping[devMapIdx['LRU']].value)
		if(lru in lrus):
			try:
				port_uut = int(mapping[devMapIdx['SW_PORT_UUT']].value)
				port_mon = int(mapping[devMapIdx['SW_PORT_SIM']].value)
				vlan_lst.append([port_uut, port_mon, lru])
			except:
				print("%s not connected in device map"%(lru))
				break
	
	#generate SW CFG
	genSwCfg(data, vlan_lst, os.path.join(outputdir, "cliList.xsf"))
	
	return 0


if __name__ == '__main__':
	sys.exit(main(sys.argv[1:]))