#$Id:$
# Module for generation of required ADS CFG files for Virtual Integration
# Platform (MVDS, SDIB) depending on XLS ICD files
#
import xlrd
from openpyxl import Workbook
import sys
import os.path
import string
import iomGenCommon
import iomGenReadDD
import iomGenJoinDD
from iomGenAds2Data import CvtCollector
from iomGenAds2Data import IOMAds2Template
from iomGenAds2Data import CVTConfigTemplate
from Cheetah.Template import Template



def checkRows(dst, src, sheet): #dst/src XL R/W cell objects lists
	result = 1 # 1 - equal
	#check NPA
	#print "check row " + repr(sheet.ncols)
	#lru_name = src[iomGenCommon.getColumnIndex(sheet)["Lru"]]
	#for pha in ["HF_IDU", "FDAS", "SYNOPT"]:
	#	if pha in lru_name:
	#		return 	1 # we don't need this
	try:
		idxPortId = iomGenCommon.getColumnIndex(sheet)["PortId"]
	except:
		idxPortId = -1
		
	for i in range(0, len(dst)):
		if i == idxPortId:
			continue
		if src[i] == dst[i]:
			continue
		elif i > 30 or (src[i] in ["HF_IDU", "FDAS", "SYNOPT"] and "Input" in sheet.sheet_name): # max columns
			result = 1
			break
		else:
			result = 0
			break
	
	return result

def mergeSheets(dst, src): #dst/src XL R/W sheet objects

	for ridx_src in range(src.nrows):
		#print str(ridx_src)
		src_row = src.row_values(ridx_src)
		found = 0
		for dst_row in dst.values():
			if(checkRows(dst_row, src_row, src) == 1):
				found = 1
				break
		
		if not found:
			dst[len(dst.keys())] = src_row		

def genMergedIcd(platform, outputdir, inputfiles):
	_icd_ = {}

	for inputfile in inputfiles:
		offset = 0
		wb = xlrd.open_workbook(inputfile)
		print "check "+inputfile
		for sheet_name in wb.sheet_names():
			if sheet_name not in _icd_.keys():
				_icd_[sheet_name] = {}
				print "add new sheet "+sheet_name
			s_i = _icd_[sheet_name]
			s_o = wb.sheet_by_name(sheet_name) 
			mergeSheets(s_i, s_o)
	
	_icd = Workbook()
	for sheet_name in _icd_.keys():
		sheet = _icd.create_sheet()
		sheet.title = sheet_name
		for key in _icd_[sheet_name].keys():
			val = _icd_[sheet_name][key]
			for i in range(len(val)):
				sheet.cell(row = key+1, column = i+1).value = val[i]
	xls_name = os.path.join(outputdir, platform + ".xlsx")
	_icd.save(xls_name)
	return xls_name

def genSDIB(modelList, outputdir):
	cvts = CvtCollector("SDIB")
	
	model = iomGenCommon.Bunch(afdx = modelList[0].afdx, can=modelList[0].can, paramXref=modelList[0].parameters, cvts=cvts)
	tmpl = Template(IOMAds2Template, searchList=[{ "model"  : model, "direction" : "output"}])
	s = tmpl.respond()
	outfn = os.path.join(outputdir,"SDIB_OUT.iom")
	open(outfn, "w").write(s)
	model = iomGenCommon.Bunch(afdx = modelList[1].afdx, can=modelList[1].can, paramXref=modelList[1].parameters, cvts=cvts)
	tmpl = Template(IOMAds2Template, searchList=[{ "model"  : model, "direction" : "input"}])
	open('_cheetadump.py', 'w').write(tmpl.generatedModuleCode())
	s = tmpl.respond()
	outfn = os.path.join(outputdir,"SDIB_IN.iom")
	open(outfn, "w").write(s)

def genMVDS(modelList, outputdir, infile):
	class_ = os.path.basename(infile).split('.')[0]
	class_ = class_.replace("-icd", "")

	cvts = CvtCollector(class_)

	# generating ADS2 I/O map and CVT for RX/TX
	model = iomGenCommon.Bunch(afdx = modelList[0].afdx, cvts=cvts)
	tmpl = Template(IOMAds2Template, searchList=[{ "model"  : model, "direction" : "input"}])
	s = tmpl.respond()
	outfn = os.path.join(outputdir, class_ + '_IN.iom')
	open(outfn, "w").write(s)
	model = iomGenCommon.Bunch(afdx = modelList[1].afdx, cvts=cvts)
	tmpl = Template(IOMAds2Template, searchList=[{ "model"  : model, "direction" : "output"}])
	s = tmpl.respond()
	
	outfn = os.path.join(outputdir, class_ + '_OUT.iom')
	open(outfn, "w").write(s)

	model = iomGenCommon.Bunch(cvtpoints=cvts.cvtpoints.values())
	tmpl = Template(CVTConfigTemplate, searchList=[{ "model"  : model}])
	s = tmpl.respond()
	outfn = os.path.join(outputdir, class_ + '.cvt')
	open(outfn, "w").write(s)
	 
def main(args):

	if len(args) <= 4:
		sys.stderr.write("Usage: iomGenVIAds2 mvds|sdib outputdir inputfile1 inputfile2 ... / ddfile1 ddfile 2 ...")
		sys.exit(1)
	
	platform = args[0]
	outputdir = args[1]
	inputfiles = args[2:]
	icdfiles = []
	ddmfiles = []
	icd_dd = 0
	for i in inputfiles:
		if i == '/':
			icd_dd = 1
		if icd_dd == 0:
			if (len(i) > 3):
				icdfiles.append(i)
		else:
			if (len(i) > 3):
				ddmfiles.append(i)

	print "platform: " + platform
	print "outputdir: " + outputdir
	print "icdfiles: " + repr(icdfiles)
	print "ddmfiles: " + repr(ddmfiles)
	
	if platform == "sdib":
		xls_name = genMergedIcd(platform, outputdir, icdfiles)
		outfn = xls_name.replace(".xlsx", "_.xlsx")
		iomGenJoinDD.main([outfn, xls_name]+ddmfiles)

		modelIn = iomGenCommon.MapReader(outfn, msgonly=False, direction='input', ignoreSrc=True)
		modelOut = iomGenCommon.MapReader(outfn, msgonly=False, direction='output', ignoreSrc=True)
		genSDIB([modelIn, modelOut], outputdir)		
		return 0
		
	if platform == "mvds": 
		for file in inputfiles:
			if not ("IDU" in file):
				modelIn = iomGenCommon.MapReader(file, msgonly=True, direction='input')
				modelOut = iomGenCommon.MapReader(file, msgonly=True, direction='output')
				genMVDS([modelIn, modelOut], outputdir, file)		
				
	return 0


if __name__ == '__main__':
	sys.exit(main(sys.argv[1:]))