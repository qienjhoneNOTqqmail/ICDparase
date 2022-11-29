
import re 
import sys
from openpyxl.reader.excel import load_workbook


ipmatch = re.compile("([ ,.<]|^)([io]p)([A-Z]|<[a-zA-Z]*>|\[a-zA-Z,|]*\])([a-zA-Z1-9]*(<[a-zA-Z]*>|\[a-zA-Z,|]*\])?[a-zA-Z1-9]*)")

def getMatchSet(text):
    res = set()
    for t in re.findall(ipmatch, text):
        res.add(t[1]+t[2]+t[3])
    return res
    
def genParamList(filename):
    wb=load_workbook(filename, read_only=False, data_only=False)
    ws = wb.get_sheet_by_name(wb.sheetnames[0])

    ws.rows[0][3].value = "Parameter Reference List"

    for row in ws.rows[1:]:
        text = row[2].value
        params = getMatchSet(text)
        s = ';\n'.join(params) 
        if len(params) != 0:
            s+= ';\n'
            
        row[3].value = s
            

    wb.save(filename)
    return 0

def testit():
    print getMatchSet("asdasd ip<a>xx asasd")
    print getMatchSet("asdasd ipA<a>xx asasd")
    print getMatchSet("asdasd  ipAyy<a>xx asasd")
    print getMatchSet("asdasd ipA[1,2]xx asasd")
    print getMatchSet("asdasd ipAxx asasd")
    print getMatchSet("asdasd ipA[1,2]xx <op<a>asdasd> asasd")

if __name__ == "__main__":
    if len(sys.argv) > 1:
        for fn in sys.argv[1:]:
            genParamList(fn)
    else:
        testit()
