'''
Convert an Excel File into one CVS file per sheet
'''

from openpyxl.reader.excel import load_workbook
import os.path

def excelToCsv(filename):
    wb = load_workbook(filename, read_only=True, data_only=True)
    
    for sheet in wb.sheetnames:
        ws = wb.get_sheet_by_name(sheet)
        outfn = os.path.splitext(filename)[0] + '-' + sheet + '.csv'
        outfile = open(outfn, "w")
        for row in ws.iter_rows():
            line = ','.join('"' + ("" if cell.value is None else str(cell.value)) + '"' for cell in row)
            outfile.write(line + '\n')
        
def main(argv):
    if len(argv) == 2:
        excelToCsv(argv[1])
    else:
        print "Usage: excelToCsv filename"
    return 0

        
if __name__ == "__main__":
    import sys
    sys.exit(main(sys.argv))        
        