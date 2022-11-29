import sys
import os
import getopt
from openpyxl import load_workbook, worksheet, Workbook
import iomGenCommon

class ads2GenDevice(object):
	''' 
		Provides different functions for calculation of ADS2 device assignments for ADS2 IOM and Switch Configurations
		Processes a given device mapping table and Excel ICD to calculate a defined ADS2 device asssignment or SW con-
		figuration
    '''
	def __init__(self, excelicd=None, devmap=None, outfile=None):
		self.devMapIdx={"LRU":-1,"LRU_PHYS_PORT":-1,"ADS_DEV_CH":-1,"PATCH_PORT_SIM":-1,"SW_PORT_SIM":-1,"SW_PORT_UUT":-1,"X_PORT":-1}
		self.excelIcd = []
		if(excelicd):
		    if(os.path.isfile(excelicd)):
		        inCfg = iomGenCommon.MapReader(excelicd, msgonly=True, direction='input', ignoreSrc=True)
		        outCfg = iomGenCommon.MapReader(excelicd, msgonly=True, direction='output', ignoreSrc=True)
		        self.excelIcd=[inCfg, outCfg]
		    else:
		        print("%s does not exists"%(excelicd))

	
		wb = load_workbook(devmap)
		self.devMap = wb.get_sheet_by_name(name = 'DEVICE_MAP')
		for row in self.devMap.rows:
			for cell in row:
				for k in self.devMapIdx.keys():
					if self.devMapIdx[k] != -1:
						continue
					else:
						if str(cell.value).startswith(k):
							self.devMapIdx[k] = ord(cell.column)-65
							break
			if -1 not in self.devMapIdx.values():
				break

		self.outFile=outfile

	def getSwPort(self, lruName, dir):
		if dir == True: #Input
		    colIdx='SW_PORT_UUT'
		else:
		    colIdx='SW_PORT_SIM'

		if lruName is not None:
		    for mapping in self.devMap.iter_rows(row_offset=1):
		        if(str(mapping[self.devMapIdx['LRU_PHYS_PORT']].value) == lruName+".afdx"):
		            try:
		                return int(mapping[self.devMapIdx[colIdx]].value)
		            except:
		                print("%s - %s not connected in device map"%(lruName, colIdx))
		                break
		    return -1
		else:
			return -1

	def getSwConfig(self):
		sw_cfg = ""
		vlsPorts = {}

		for dir in range(len(self.excelIcd)):
		    for vl in self.excelIcd[dir].afdx.vls.values():
		        port = self.getSwPort(vl.lru, dir==0)
		        if port < 1:
		            continue
		        if vl.macAdd not in vlsPorts.keys():
		            vlsPorts[vl.macAdd] = []
		        if not str(port) in vlsPorts[vl.macAdd]:
		            vlsPorts[vl.macAdd].append(str(port))
		
		for vl in vlsPorts.keys():
		    sw_cfg+="Create fdbentry "+vl+" vlan vafdx ports %s\n"%(( ",".join(vlsPorts[vl])))	
		if(self.outFile is None):
			print sw_cfg
		else:
			open(self.outFile, "w").write(sw_cfg)
	
	def getAdsDevice(self, physPort=None):
		if physPort is not None:
			for mapping in self.devMap.iter_rows(row_offset=1):
				if(str(mapping[self.devMapIdx['LRU_PHYS_PORT']].value) == physPort):
					return str(mapping[self.devMapIdx['ADS_DEV_CH']].value)
			return "ADS-DEV-TBD"
		else:
			return "ADS-DEV-TBD"

	def getAdsDeviceList(self, lru):
		devList = []
		if lru is not None:
		    for mapping in self.devMap.iter_rows(row_offset=1):
		        if(str(mapping[self.devMapIdx['LRU']].value) == lru):
		            if str(mapping[self.devMapIdx['ADS_DEV_CH']].value) not in devList:
		                devList.append(str(mapping[self.devMapIdx['ADS_DEV_CH']].value))
		return devList

	def setIcdMap(self, icdMap):
           self.excelIcd=icdMap
def usage():
	print 'Usage: ads2GenDeviceMapping -i excelicd -m devicemapping -o outputfile'
	
def main():
	# Parse command line arguments
	outFile		= None ## full filepathname of SW CFG to be created
	excelIcd	= None ## full filepathname of the Excel ICD to be considered
	devMap		= None ## full filepathname of the Device Mapping to be considered Excel Sheet "DEVICE_MAP"
						## columns: LRU_PHYS_PORT|ADS_DEV_CH|PATCH_PORT_SIM|SW_PORT_SIM|SW_PORT_UUT|X_PORT

	opts, args = getopt.getopt(sys.argv[1:], 'hm:i:o:', ['excelicd=','devmap=', 'outfile=','help'])
	
	for o, a in opts:
		if o in ['-h', '--help']:
			usage()
		elif o in ['-o', '--outfile']:
			outFile = a
		elif o in ['-m', '--devmap']:
			devMap = a
		elif o in ['-i', '--excelicd']:
			excelIcd = a
		else:
			usage()
			return -1
	if (outFile is None) or (excelIcd is None) or (devMap is None): 
		usage()
		return -1
		
	genCfg = ads2GenDevice(excelIcd, devMap, outFile)
	genCfg.getSwConfig()
	return 0

if __name__ == "__main__":
	sys.exit(main())

