'''
 perform merge between:
 - The old AsUsed Module
 - The new AsReceived Module
 
according to the control in diff3 output

Open asUsed Module in RdWr mode
Apply all changes from all sheets

'''

from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.styles import Color, Font, Side, Border, PatternFill, Style
from openpyxl.reader.excel import load_workbook
from collections import OrderedDict

yellow   = Color(rgb='00ffff00')
red      = Color(rgb='00ff0000')
blue     = Color(rgb='00b8cce4')

thinline = Side(style='thin')
border   = Border(left=thinline, right=thinline, top=thinline, bottom=thinline)

hdrstyle = Style(
    font=Font(b=True),
    fill=PatternFill(patternType='solid', bgColor=yellow, fgColor=yellow),
    border=border
)

changedstyle = Style(
    fill=PatternFill(patternType='solid', bgColor=red, fgColor=red),
    border=border
)

emptystyle = Style(
    fill=PatternFill(patternType='solid', bgColor=blue, fgColor=blue),
    border=border
)

normstyle=Style(border=border)



class DiffRecord:
    def __init__(self, asUsed, asRcvdOld, asRcvdNew, diff):
        self.asUsed = asUsed
        self.asRcvdOld = asRcvdOld
        self.asRcvdNew = asRcvdNew
        self.diff = diff

    
class IcdDiff:

    KEYCOLUMNS = {
        # sheetname: keycolumnnames
        'InputAfdxMessages':    'UniqueKey',
        'OutputAfdxMessages':   'UniqueKey',
        'InputCanMessages':     'UniqueKey',
        'OutputCanMessages':    'UniqueKey',
        'InputA429Labels':      'UniqueKey',
        'OutputA429Labels':     'UniqueKey',
        'InputSignals':         'RpName',
        'OutputSignals':        'UniqueName',
    }
    
    EXCLUDECOLUMNS = {
        'InputAfdxMessages':    ("Status", "ICDFix", "Comment", "Change History"),
        'OutputAfdxMessages':   ("Status", "ICDFix", "Comment", "Change History"),
        'InputCanMessages':     ("Status", "ICDFix", "Comment", "Change History"),
        'OutputCanMessages':    ("Status", "ICDFix", "Comment", "Change History"),
        'InputA429Labels':      ("Status", "ICDFix", "Comment", "Change History"),
        'OutputA429Labels':     ("Status", "ICDFix", "Comment", "Change History"),
        'InputSignals':         ("Status", "ICDFix", "Comment", "Change History"),
        'OutputSignals':        ("Status", "ICDFix", "Comment", "Change History"),
    }
        
    def __init__(self, reportFN, asUsedFN, asRcvdOldFN, asRcvdNewFN=None):
        self.report = Workbook()
        self.report.save(reportFN)
        
        print "Loading:", asUsedFN 
        self.asUsed    = load_workbook(asUsedFN,    read_only=False, data_only=False)
        print "Loading:", asRcvdOldFN 
        self.asRcvdOld = load_workbook(asRcvdOldFN, read_only=True, data_only=False)
        if asRcvdNewFN:
            print "Loading:", asRcvdNewFN 
            self.asRcvdNew = load_workbook(asRcvdNewFN, read_only=True, data_only=False)
        else:
            self.asRcvdNew = None

        self.sheetnames = self.asUsed.sheetnames
        
    def getHeader(self, ws):
        for row in ws.iter_rows():
            header = [str(cell.value).strip() for cell in row]
            return header
        
        return None
    
    def normval(self, value):
        val = str(value)
        if val in ("TRUE", "True", "true", "WAHR", "Wahr", "wahr"):
            val = "True"
        elif val in ("FALSE", "False", "false", "FALSCH", "Falsch", "falsch"):
            val = "False"
        return val

    def diff2lines(self, line1, line2, exclude=()):
        res = []
        for col in line1.keys():
            if not col in exclude:
                v1 = self.normval(line1[col])
                v2 = self.normval(line2[col])
                if v1 != v2:
                    res.append(col)
                
        return res
        
    def diff3lines(self, line1, line2, line3, exclude=()):
        res = []
        for col in line1.keys():
            if not col in exclude:
                v1 = self.normval(line1[col])
                v2 = self.normval(line2[col])
                v3 = self.normval(line3[col])
                if v1 != v2 or v2 != v3 or v1 != v3:
                    res.append(col)
                
        return res
    

    def makeDict(self, sheet, keycol):
        res = OrderedDict()
        rowidx = 0
        for row in sheet.iter_rows():
            if rowidx == 0:
                header = [str(cell.value).strip() for cell in row]
                rowidx += 1
            else:
                rec = {}
                colidx = 0
                for cell in row:
                    rec[header[colidx]] = cell.value
                    colidx += 1
                key = rec[keycol].strip()
                res[key] = rec
        return res
            
    def compare2Sheets(self, asUsed, asRcvdOld, keycol='', exclude=()):
        res = []
        
        asUsedDict    = self.makeDict(asUsed, keycol)
        asRcvdOldDict = self.makeDict(asRcvdOld, keycol)
        
        lineno = 0
        for key in asUsedDict.keys():
            asUsedLine    = asUsedDict.get(key)
            asRcvdOldLine = asRcvdOldDict.get(key)
            
            if asRcvdOldLine is not None:
                diff = self.diff2lines(asUsedLine, asRcvdOldLine, exclude=exclude)
                if len(diff) == 0:
                    continue;   # nothing to do, the lines are unchanged
                else:
                    res.append(DiffRecord(asUsedLine, asRcvdOldLine, None, diff))
            else:
                res.append(DiffRecord(asUsedLine, None, None, []))
                    
            
        # check for lines in asRcvdOld not existing in asUsed. 
        # These have not been considered above
        for key in asRcvdOldDict.keys():

            if asUsedDict.get(key) is not None:
                # we have already visited this, so continue
                continue

            asRcvdOldLine = asRcvdOldDict.get(key)
            res.append(DiffRecord(None, asRcvdOldLine, None, []))

        print "diffrecords:", len(res); return res

    def compare3Sheets(self, asUsed, asRcvdOld, asRcvdNew, keycol="", exclude=()):
        res = []
        
        asUsedDict    = self.makeDict(asUsed,    keycol)
        asRcvdOldDict = self.makeDict(asRcvdOld, keycol)
        asRcvdNewDict = self.makeDict(asRcvdNew, keycol)
        
        lineno = 0
        for key in asUsedDict.keys():
            asUsedLine    = asUsedDict.get(key)
            asRcvdOldLine = asRcvdOldDict.get(key)
            asRcvdNewLine = asRcvdNewDict.get(key)
            
            # consider all cases:
            # X     Y      Z
            # X     Y     None 
            # X    None    Z 
            # X    None   None 

            if asRcvdOldLine is not None and asRcvdNewLine is not None:
                diff = self.diff3lines(asUsedLine, asRcvdOldLine, asRcvdNewLine, exclude=exclude)
                if len(diff) == 0:
                    continue;   # nothing to do, the lines are unchanged
                else:
                    res.append(DiffRecord(asUsedLine, asRcvdOldLine, asRcvdNewLine, diff))
            elif asRcvdOldLine is not None and asRcvdNewLine is None:
                diff = self.diff2lines(asUsedLine, asRcvdOldLine, exclude=exclude)
                res.append(DiffRecord(asUsedLine, asRcvdOldLine, None, diff))
            elif asRcvdOldLine is None and asRcvdNewLine is not None:
                diff = self.diff2lines(asUsedLine, asRcvdNewLine, exclude=exclude)
                res.append(DiffRecord(asUsedLine, None, asRcvdNewLine, diff))
            else:
                res.append(DiffRecord(asUsedLine, None, None, []))
                    
            
        # check for lines in asRcvdOld not existing in asUsed. 
        # These have not been considered above
        for key in asRcvdOldDict.keys():

            if asUsedDict.get(key) is not None:
                # we have already visited this, so continue
                continue
            
            # consider all cases:
            # None      Y      Z
            # None      Y     None

            asRcvdOldLine = asRcvdOldDict.get(key)
            asRcvdNewLine = asRcvdNewDict.get(key)

            if asRcvdNewLine is not None:
                diff = self.diff2lines(asRcvdOldLine, asRcvdNewLine, exclude=exclude)
                res.append(DiffRecord(None, asRcvdOldLine, asRcvdNewLine, diff))
            else:
                res.append(DiffRecord(None, asRcvdOldLine, None, []))

        # check for lines in asRcvdNew not existing in asUsed and asRcvdOld. 
        # These have not been considered above
        for key in asRcvdNewDict.keys():
            
            if asUsedDict.get(key) is not None:
                # we have already visited this, so continue
                continue

            if asRcvdOldDict.get(key) is not None:
                # we have already visited this, so continue
                continue

            res.append(DiffRecord(None, None, asRcvdNewLine, []))
        
        return res
    
    def setcolors(self, row, difflist):
        colidx = 0
        rowempty = False

        for cell in row:
            if colidx < 2:
                cell.style = normstyle
            elif colidx == 2:
                if cell.value == 'E':
                    rowempty = True
                    cell.style = emptystyle
                elif difflist:
                    cell.style = changedstyle
                else:
                    cell.style = normstyle
            else:
                if rowempty:
                    cell.style = emptystyle
                elif colidx - 3 in difflist:
                    cell.style = changedstyle
                else:
                    cell.style = normstyle
            colidx += 1
            

    
    def formatReportSheet(self, sheetname, header, diffData, linenames):
        '''
        create a formatted output sheet from the diff data
        and add it to the report workbook
        '''

        refSheet = self.asUsed.get_sheet_by_name(sheetname)
        repSheet = self.report.create_sheet(title=sheetname)
        repHeader = ['ICD', 'Apply', 'Compare'] + header
        repSheet.append(repHeader)
        
        col2idx = {}
        for idx in range(len(header)):
            col2idx[header[idx]] = idx

        # set column width
        for colidx in range(1, len(header)+1):
            letterIn  = get_column_letter(colidx)
            letterOut = get_column_letter(colidx + 3)
            repSheet.column_dimensions[letterOut].width = \
                refSheet.column_dimensions[letterIn].width
            
        
        # set header style
        row = repSheet.rows[0]
        idx = 0
        for cell in row:
            if idx < 3:
                cell.style = normstyle
            else:
                cell.style = hdrstyle
            idx += 1
        
        for diffItem in diffData:
            if diffItem.diff:
                code = "C"
            else:
                code = ""
            
            if diffItem.asUsed:
                
                line = [linenames[0], '', code] + [diffItem.asUsed[col] for col in header] 
            else:
                line = [linenames[0], '', 'E']
            repSheet.append(line)
                
            if diffItem.asRcvdOld:
                line = [linenames[1], '', code] + [diffItem.asRcvdOld[col] for col in header]
            else:
                line = [linenames[1], '', 'E']
            repSheet.append(line)

            if len(linenames) == 3:
                if diffItem.asRcvdNew:
                    line = [linenames[2], '', code] + [diffItem.asRcvdNew[col] for col in header]
                else:
                    line = [linenames[2], '', 'E']
                repSheet.append(line)
            repSheet.append([]); 
            
        count = 0
        for row in repSheet.rows[1:]:
            r0v = row[0].value
            if r0v:
                if r0v == linenames[0]:
                    diffitem = diffData[count]
                    difflist = [col2idx[col] for col in diffitem.diff]
                    count += 1
                self.setcolors(row, difflist)


    def compare3(self):
        '''
        Compare all sheets of the three excel files 
        '''

        for sheet in self.sheetnames:
            print "Comparing:", sheet
            s1 = self.asUsed.get_sheet_by_name(sheet)
            s2 = self.asRcvdOld.get_sheet_by_name(sheet)
            s3 = self.asRcvdNew.get_sheet_by_name(sheet)
            
            keyname = self.KEYCOLUMNS[sheet]
            excludes = self.EXCLUDECOLUMNS[sheet]
                
            diffdata = self.compare3Sheets(s1, s2, s3, keycol=keyname, exclude=excludes)
            if diffdata:
                header = self.getHeader(s1)
                self.formatReportSheet(sheet, header, diffdata, ("CLEAN", "OLD", "NEW"))
        print "Done"

    def compare2(self):
        '''
        Compare all sheets of the two excel files 
        '''
        
        for sheet in self.sheetnames:
            print "Comparing:", sheet
            s1 = self.asUsed.get_sheet_by_name(sheet)
            s2 = self.asRcvdOld.get_sheet_by_name(sheet)
            
            keyname  = self.KEYCOLUMNS[sheet]
            excludes = self.EXCLUDECOLUMNS[sheet]

            diffdata = self.compare2Sheets(s1, s2, keycol=keyname, exclude=excludes)
            if diffdata:
                header = self.getHeader(s1)
                self.formatReportSheet(sheet, header, diffdata, ("CLEAN", "NEW"))

        print "Done"
        
    def save(self, filename):
        if len(self.report.worksheets) > 1:
            del self.report.worksheets[0]

        self.report.save(filename)
        
def main(argv):
    
    if len(argv) == 5:
        try:
            difftool = IcdDiff(argv[1], argv[2], argv[3], argv[4])  
        except Exception, e:
            print str(e)
            return -1
        difftool.compare3()
    elif len(argv) == 4:
        try:
            difftool = IcdDiff(argv[1], argv[2], argv[3])   
        except Exception, e:
            print str(e)
            return -1
        difftool.compare2()
    else:
        print "Usage: icdDiff report asUsed asReceivedOld [asReceivedNew] "
        return -1

    difftool.save(argv[1])

        
if __name__ == "__main__":
    import sys
    sys.exit(main(sys.argv))        
        
        
        
        
        
        
        
        
        
    
    