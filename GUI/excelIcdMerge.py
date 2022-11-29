'''
 perform merge between:
 - The old AsUsed Module
 - The new AsReceived Module
 
according to the control in diff3 output

Open asUsed Module in RdWr mode
Apply all changes from all sheets

'''

from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.styles import Color, Font, Side, Border, PatternFill, Style
from openpyxl.reader.excel import load_workbook

import excelToCsv


    
class IcdMerge:

    KEYCOLUMNS = {
        # sheetname: keycolumnnames
        'InputAfdxMessages':    'UniqueKey',
        'OutputAfdxMessages':   'UniqueKey',
        'InputCanMessages':     'UniqueKey',
        'OutputCanMessages':    'UniqueKey',
        'InputA429Labels':      'UniqueKey',
        'OutputA429Labels':     'UniqueKey',
        'InputSignals':         'RpName',
        'OutputSignals':        'UniqueName',
    }
        
    def __init__(self, reportFN, asUsedFN):
        
        print "Loading:", asUsedFN 
        self.asUsed    = load_workbook(asUsedFN,    read_only=False, data_only=False)
        self.sheetnames = self.asUsed.sheetnames
        
        print "Loading:", reportFN
        self.report    = load_workbook(reportFN,    read_only=True, data_only=True)

        self.newWb = Workbook()
        del self.newWb.worksheets[0]
        


    def getHeader(self, ws):
        for row in ws.iter_rows():
            header = [str(cell.value).strip() for cell in row]
            return header
        
        return None

    
    def makeDict(self, sheet, keyidx):
        res = {}
        for row in sheet.iter_rows():
            key = row[keyidx].value.strip()
            res[key] = row
        return res
            
    def appendrow(self, sheet, newWs, datarow, reprow):
        # create a row from the content of reprow and append it to asUsed
        fields = [reprow[idx].value for idx in range(3, len(reprow))]
        sheet.append(fields)
        newWs.append(fields)
                
    def updaterow(self, sheet, newWs, datarow, reprow):
        # update cells in datarow with data from reprow
        fields = [reprow[idx].value for idx in range(3, len(reprow))]
        idx = 0
        for cell in datarow:
            cell.value = fields[idx]
            idx += 1
        newWs.append(fields)
        
    def deleterow(self, sheet, newWs, datarow, reprow):
        # mark datarow as deleted
        datarow[0].value = "Deleted"
        fields = [cell.value for cell in datarow]
        fields[0] = "Deleted"
        newWs.append(fields)
        
        
    
        
    def mergeSheet(self, asUsed, report, keyidx=0, header=None, sheetname=""):
        
        newWs = self.newWb.create_sheet(title=sheetname)
        newWs.append(header)
        
        asUsedDict    = self.makeDict(asUsed, keyidx)
        
        lineno = 0
        for reportLine in report.iter_rows():
            if reportLine[0].value == 'CLEAN':
                repAsUsed = reportLine
            elif reportLine[0].value == 'OLD':
                repOld = reportLine
            elif reportLine[0].value == 'NEW':
                repNew = reportLine
                # check for command in Apply column. If is is not "A" or "Y", skip the block
                if repNew[1].value in ('Y', 'y', 'Yes', 'yes', 'T', 't', 'true', 'True'):
                
                    # OK. We have to apply this record 
                    #
                    # consider following case:
                    # "New" line is marked "C", "asUsed" line is marked "C":
                    #    copy New line to asUsed [New line key] and delete status 
                    # 
                    # "New" line is marked "C", "asUsed" line is marked "E":
                    #    append New line to asUsed 
                    # 
                    # "New" line is marked "E", "asUsed" line is marked "C":
                    #    set Status of asUsed [asUsed line key] to "Deleted"
                    #
                    # "New" line is marked "E", "asUsed" line is marked "E":
                    #    do nothing
        
                    if repAsUsed[2].value == 'C' and repNew[2].value == 'C':
                        key = repNew[keyidx + 3].value
                        updateRow = asUsedDict.get(key)
                        self.updaterow(asUsed, newWs, updateRow, repNew)
                    elif repAsUsed[2].value == 'E' and repNew[2].value != 'E'  :
                        self.appendrow(asUsed, newWs, updateRow, repNew)
                    elif repAsUsed[2].value != 'E' and repNew[2].value == 'E':
                        key = repAsUsed[keyidx + 3].value
                        updateRow = asUsedDict.get(key)
                        self.deleterow(asUsed, newWs, updateRow, repNew)
                    elif repAsUsed[2].value == 'E' and repNew[2].value == 'E':
                        pass
    
                

    def mergeBook(self):
        '''
        Merge all sheets of an excel ICD
        '''
        
        for sheet in self.sheetnames:
            s1 = self.asUsed.get_sheet_by_name(sheet)
            s2 = self.report.get_sheet_by_name(sheet)
            if s2:
                print "Merging:", sheet          
                header = self.getHeader(s1)
                keyname = self.KEYCOLUMNS[sheet]
                keyidx = 0
                for idx in range(len(header)):
                    if header[idx] == keyname:
                        keyidx = idx
                        break
    
                self.mergeSheet(s1, s2, keyidx=keyidx, header=header, sheetname=sheet)
                
        print "Done"
        
    def save(self, asUsedFN, newWbFN):
        self.newWb.save(newWbFN)
        self.asUsed.save(asUsedFN)
        
def main(argv):
    #try:
        if len(argv) >= 3:
            mergetool = IcdMerge(argv[1], argv[2])  
            mergetool.mergeBook()
        else:
            print "Usage: icdMerge report asUsed [output] "
            return -1        if len(argv) == 4:
            outfn = argv[3]
        else:
            outfn = "new_" + argv[2] 
            
        mergetool.save(argv[2], outfn)
		
        excelToCsv.excelToCsv(outfn)

        return 0

#    except Exception, e:
#        sys.stderr.write(str(e))
#        return 1

        
if __name__ == "__main__":
    import sys
    sys.exit(main(sys.argv))        
        
        
        
        
        
        
        
        
        
    
    