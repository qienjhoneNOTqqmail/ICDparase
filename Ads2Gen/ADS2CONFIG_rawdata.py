import copy
import re
import os
import sys
sys.path.append("..")
import getopt
from Common.bunch import Bunch
import Common.logger as logger
import xlrd
from collections import defaultdict
import Cheetah.Template 
from ssl import DefaultVerifyPaths
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.styles import Color, Font, Side, Border, PatternFill, Style
from openpyxl.reader.excel import load_workbook
from string import Template
import AnalysisExcel
from XmlImport.imtExcelRW import  genExcelFile



Aliases = []

TEMPLATE='Create fdbentry $mac vlan vafdx ports $portlist\n'
def Generate_Fanout_Config(outdir,devmap,inputvl,outputvl):
    
    if devmap is None:
         return
    filea=open(os.path.join(outdir,'SSDL_FANOUT_Conf5.1.xsf'),'w')
    
    MessagePort = Workbook()
    ws = MessagePort.create_sheet(title='Message-Port')
    ws.append(['VLID','LRU','Message','Destination UDP Port','Fanout Portlist'])
    # write header
    
    temp = Template(TEMPLATE)
    excludeid = []
    for vlid, item in inputvl.items():
        portlist=[]
        msgport = []
        msg=[]
        udpport = []
        for receivelru in item.receivelru:
            #print receivelru
            port = devmap.Map.get(receivelru)[0]['FanPort']
            if int(vlid) == 33211:
                print "#############&&&&&&&&&&&&&&&"+str(port)
            portlist.append(port)
        if outputvl.get(vlid) is not None:
            monitorport = devmap.Map.get(item.lru)[0]['swportsim']
            portlist.append(monitorport)
            if int(monitorport) == 8:
                portlist.append(9)
            excludeid.append(vlid)
        portlist = list(set(portlist))
        portlist = ','.join(str(i) for i in portlist)

        filea.write(temp.substitute({'mac': item.macaddr,                                     
                                     'portlist':portlist,
                                     }))
        for mp in item.msgport:
            #print str(mp[0])+'^^^^^^^^^^'+str(mp[1])
            msg.append(mp[0])
            udpport.append(str(mp[1]))
        msgport = [vlid,item.lru,'\n'.join(msg),'\n'.join(udpport),portlist]
        ws.append(msgport)
    for vlid, item in outputvl.items():
        if vlid in excludeid:
            continue
        msg=[]
        udpport = []
        
        port1 = devmap.Map.get(item.lru)[0]['swportsim']
        if int(port1) == 8:
            porttem=[port1,9]
        else:
            porttem=[port1]
        port = ','.join(str(i) for i in porttem)
        
        filea.write(temp.substitute({'mac': item.macaddr,                                     
                                     'portlist':port,
                                     }))
        for mp in item.msgport:
            msg.append(mp[0])
            udpport.append(str(mp[1]))
        msgport = [vlid,item.lru,'\n'.join(msg),'\n'.join(udpport),port]
        ws.append(msgport)
    filea.close() 
    MessagePort.save(os.path.join(outdir,'Message-Ports5.1.xlsx'))
 

def Generate_Fanout_Config_52(outdir,devmap,inputvl,outputvl):
        
    if devmap is None:
        return
    filea=open(os.path.join(outdir,'SSDL_FANOUT_Conf5.2.xsf'),'w')
    
    MessagePort = Workbook()
    ws = MessagePort.create_sheet(title='Message-Port')
    ws.append(['VLID','LRU','Message','Destination UDP Port','Fanout Portlist'])
    # write header
    
    temp = Template(TEMPLATE)
    #excludeid = []
    #for item in inputvl.values()+outputvl.values():
    #    print item.lru + str(item.vlid)
    #print "input vl : %d\n" % len(inputvl)
    #print "output vl : %d\n" % len(outputvl)
    vllrus= dict()
    
    for vlid,item in inputvl.items():  # inputvl has contained the vlid that inside the DS system
        if item.lru not in vllrus:
            vllrus[item.lru] = 1
        else:
            vllrus[item.lru] += 1
    #for key, val in vllrus.items():
    #    print "%s:  %d"%(key,val)
    for vlid, item in inputvl.items():
        portlist=[]
        msgport = []
        txlru = item.lru
        msg = []
        udpport = []
        #print "txlru=%s"%txlru
        #if not txlru:
        #    print "find it"
        #    raw_input()
        
        #print "devmap.Map.get(txlru)[0]=%s \n"%devmap.Map.get(txlru)[0]
        #raw_input()
        
        port = devmap.Map.get(txlru)[0]['FanPort']
        portlist.append(port)

        portlist = list(set(portlist))
        portlist = ','.join(str(i) for i in portlist)

        filea.write(temp.substitute({'mac': item.macaddr,                                     
                                     'portlist':portlist,
                                     }))
        for mp in item.msgport:
            msg.append(mp[0])
            udpport.append(str(mp[1]))
        msgport = [vlid,item.lru,'\n'.join(msg),'\n'.join(udpport),portlist]
        ws.append(msgport)
    #for vlid, item in outputvl.items():
    #    filea.write(temp.substitute({'mac': item.macaddr,                                     
    #                                 'portlist':'8',
    #                                 }))
    #    msgport = [vlid,item.lru,item.msgname,item.destupd,'8']
    #    ws.append(msgport)
    filea.close() 
    MessagePort.save(os.path.join(outdir,'Message-Ports5.2.xlsx'))             

class ADS2_Device_Map(object):
    def __init__(self,file):
        wb = xlrd.open_workbook(file)
        self.ws = wb.sheet_by_name('DEVICE_MAP')
        
        self.Map = dict()
        
        cdix = AnalysisExcel.Get_Column_Index(self.ws)
        
                
        for rowid in range(1,self.ws.nrows):
            data = dict()
            row = self.ws.row(rowid)
            
            data["Lru"] = sval(row,cdix['LRU'])
            data['Lru_Channel'] = sval(row,cdix['LRU_PHYS_PORT'])
            data['SSDL_Device'] = sval(row,cdix['ADS_DEV_CH'])
            data['DeviceType'] = sval(row,cdix['DEVICE_TYPE'])
            data['pathpanelport'] = ival(row, cdix['PATCH_PORT_SIM'])
            data['swportsim'] = ival(row,cdix['SW_PORT_SIM'])
            data['FanPort'] = ival(row,cdix['SW_PORT_UUT'])
            data['realsim'] = sval(row,cdix['SIM/REAL'])
            
            key = data['Lru']

            if key not in self.Map:
                self.Map[key] = [data]  # use list for that CAN bus type has A and B channel, but their keys are same
            else:
                item = self.Map[key]
                item.append(data)
        
        

def Get_Column_Index(sheet):
    cid=dict()
    for col in range(sheet.ncols):
        key = sval(sheet.row(0),col)
        cid[key]= col
    return cid

def ival(r, col, base= 10):
    f = r[col].value
    if type(f) == type(1.0):
        return int(r[col].value)
    elif type(f) == type(1):
        return f
    else:
        s = f.encode('utf-8').strip()
        if s =='':
            return None
        if s.startswith("0x"):
            return int(s[2:], 16)
        elif s.endswith("b"):
            return int(s[0:-1], 2)
        else: 
            return int(s, base)
def sval(row, index):
    return row[index].value.encode("UTF-8").strip() 

def fval(row, index):
    val = row[index].value
    if val:
        return float(row[index].value)
    else:
        return None

def bval(row, index):
    
    val = row[index].value
    if type(val) == type(1.0) or type(val) == type(1):
        return bool(val)
    else:
        s=val.encode('UTF-8').strip()    
        s = s.lower()
        if s in ['true','y','yes']:
            return True
        elif s in ['false','n','no']:
            return False
        else:
            return False         

def str_to_int(val, base = 10):
    if val:
        if type(val) == type(1.0) or type(val) == type(1):
            return int(val)
        else:
            s = val.encode('utf-8').strip()
            if s=='NA':
                return None
            if s =='':
                return None
            if s.startswith("0x"):
                return int(s[2:], 16)
            elif s.endswith("b"):
                return int(s[0:-1], 2)
            else: 
                if '.' in s:
                    a = float(s)
                    return int(a)
                else:
                    return int(s, base)

def str_to_float(val):    
    if val:
        return float(val)
    else:
        return None
    
def str_to_bool(val):
    if type(True) == type(val):
        return val
    else:
        s=val.encode('UTF-8').strip()    
        s = s.lower()
        if s in ['true','y','yes']:
            return True
        elif s in ['false','n','no']:
            return False
        else:
            return False
        

def _getdictvalue(dict,key, cast = None):
    
    value = None
    if dict.has_key(key):
        value = dict.get(key)
        if cast:
            if cast == "INT":
                return str_to_int(value)
            elif cast == "FLOAT":
                return str_to_float(value)
            elif cast == "BOOL":
                return str_to_bool(value)
            else:
                return None
        else:
            return value
    else:
        logger.error('has no key %s' % key)
        return None
        
        


ONEBIT_TYPES   = set(("BOOL","DIS"))
FULLBYTE_TYPES = set(("INT", "UINT","SINT","CHAR", "FLOAT" ))
INTEGER_TYPES  = set(("INT", "UINT"))
HALF_INTEGER_TYPES = set(("SINT",))
BLK_TYPES =set(("OPAQUE","BLK"))

DS_Internal_Lru = ['HF_IDURIGHTINBOARD', 'FDAS_L1',  'HF_IDUCENTER', 'SynopticMenuApp_R', 'SYNOPTICMENUAPP_R', 'HF_IDULEFTINBOARD', 
                   'HF_IDURIGHTOUTBOARD', 'SynopticPageApp_R','SYNOPTICPAGEAPP_R',  'FDAS_R3', 'HF_IDULEFTOUTBOARD', 'FDAS_L3',  'IMA_DM_L4',  'IMA_DM_R4',  'IMA_DM_L5', 'SynopticPageApp_L', 'SYNOPTICPAGEAPP_L', 
                   'SynopticMenuApp_L','SYNOPTICMENUAPP_L']

SSDL_1 = {"Name":'SSDL#1',
          "AFDX":['FSIB_AFDX_01','FSIB_AFDX_02','FSIB_AFDX_03','FSIB_AFDX_04','FSIB_AFDX_05','FSIB_AFDX_06','FSIB_AFDX_07','FSIB_AFDX_08'],
          'DPC':['FSIB-DPC-01-01','FSIB-DPC-01-02','FSIB-DPC-01-03']}

SSDL_2 = {"Name":'SSDL#2',
          "AFDX":['FSIB_AFDX_01','FSIB_AFDX_02','FSIB_AFDX_03','FSIB_AFDX_04','FSIB_AFDX_05','FSIB_AFDX_06','FSIB_AFDX_07','FSIB_AFDX_08'],
          'DPC':['FSIB-DPC-01-01','FSIB-DPC-01-02','FSIB-DPC-01-03']}

LRU_TO_BOARD = {'FSIB_AFDX_01' :   ['HF_FCM_1','HF_FCM_2','HF_FCM_3'], #'NTM_L1', 'NTM_R3','NTM_L4', 
                'FSIB_AFDX_02' :   ['RGW04_NonA664_In', 'RGW02_NonA664_In','RGW06_NonA664_In','RGW03_NonA664_In','RGW01_NonA664_In','RGW05_NonA664_In'  ],
                'FSIB_AFDX_03' :   ['RGW07_NonA664_In','RGW08_NonA664_In', 'RGW13_NonA664_In', 'HF_AHMUINSTANCE'],
                'FSIB_AFDX_04' :   ['RGW09_NonA664_In', 'RGW10_NonA664_In','RGW11_NonA664_In', 'RGW14_NonA664_In','RGW16_NonA664_In','RGW12_NonA664_In','RGW15_NonA664_In'],
                'FSIB_AFDX_05' :   ['HF_FADEC_R_CHA', 'HF_FADEC_R_CHB','HF_FADEC_L_CHA', 'HF_FADEC_L_CHB'],
                'FSIB_AFDX_06' :   ['HF_RPDU_UP_2B','HF_RPDU_UP_2A',  'HF_RPDU_UP_1B','HF_RPDU_UP_1A', ],
                'FSIB_AFDX_07' :   ['HF_ISSPROCESSINGUNIT_R','HF_ISSPROCESSINGUNIT_L',  'HM_L1','HM_R1','APP_FMS_CORE_1','APP_FMS_TUNE_2', 'APP_FMS_NAV_1', 'APP_FMS_CORE_2', 'APP_FMS_TUNE_1',  'APP_FMS_NAV_2',  'APP_FMS_GUIDANCE_1','APP_FMS_GUIDANCE_2','APP_FMS_DATALINK_2',  'APP_FMS_DATALINK_1']}

#this two function are no used 
def Generate_IOMConfig(inputmap,hfname,outdir,gac):
    templatedir = os.path.dirname(__file__)
    
    tmpl = Template(file= os.path.join(templatedir, "ads2genAFDXV4.tpl"), 
                searchList=[{                    
                    "model" : inputmap,                   
                    "device": 'FSIB_AFDX_01',
                    "gac" : gac,
                    }])

    fn = os.path.join(outdir, hfname + '.iom')
    temfile = tmpl.respond()
    open(fn, "w").write(temfile)
'''    
def Generate_CVTConfig(cvt):
    templatedir = os.path.dirname(__file__)
    
    tmpl = Template(file= os.path.join(templatedir, "ads2genCVT.tpl"), 
                searchList=[{                    
                    "cvtpoints" : cvt,
                    }])

    fn = os.path.join(outdir, hfname + '.cvt')
    temfile = tmpl.respond()
    open(fn, "w").write(temfile)
'''

def GenerateComponent(lrus,outdir,filename=None): #lrus is a list, that included in the cmp document, filename is the name of cmp file
    templatedir = os.path.dirname(__file__)
    
      
    tmpl = Cheetah.Template.Template(file= os.path.join(templatedir, "ads2genCMP.tpl"), 
            searchList=[{                    
                "lrus" :lrus,
                }])
    if filename:
        fn = os.path.join(outdir, filename + '.cmp')
    else:
        fn = os.path.join(outdir, lrus[0] + '.cmp')
    temfile = tmpl.respond()
    open(fn, "w").write(temfile)


def GenerateconfigDiscrete(devmap):
    templatedir = os.path.dirname(__file__)           
    #gac.cvts.clear()    
        
    lrus={}
    for key, value1 in  devmap.Map.items():
        for value in value1:
            if value['DeviceType'] == 'Discrete' and not value['SSDL_Device'].endswith('XX'):
                #print value['SSDL_Device']
                key1 = value['Lru']
                board = value['SSDL_Device'].split(',')[0]
                if key1 not in lrus:
                    lrus[key1]=[(value['Lru'],value['Lru_Channel'],board,value['SSDL_Device'].split(',')[1])]
                else:
                    item = lrus[key1]
                    item.append((value['Lru'],value['Lru_Channel'],board,value['SSDL_Device'].split(',')[1]))   
    
    
    for lruname in lrus.keys():
        cvts= CVTConnect(lruname+'_Dist')
        #if lruname not in devmap.Map:
        #    print 'Lru: %s not exist in Device Map file'% lruname
        #    continue
        device = lrus.get(lruname)
        
        board={}
        for item in device:
            if item[2] not in board:
                board[item[2]] = [(item[0],item[1],int(item[3]))]
            else:
                board[item[2]].append((item[0],item[1],int(item[3])))
        
        #for key,item in board.items():
        #    print key
        #    print item
        
        tmpl = Cheetah.Template.Template(file= os.path.join(templatedir, "ads2genDIST.tpl"), 
                searchList=[{                    
                    "device": board,
                    "cvts" : cvts,                    
                    }])
        fn = os.path.join(outdir, lruname + '_Dist.iom')
        temfile = tmpl.respond()
        open(fn, "w").write(temfile)
        
        tmpl = Cheetah.Template.Template(file= os.path.join(templatedir, "ads2genCVT.tpl"), 
                    searchList=[{                    
                        "cvtpoints" :cvts.cvts,
                        }])
    
        fn = os.path.join(outdir, lruname + '_Dist.cvt')
        temfile = tmpl.respond()
        open(fn, "w").write(temfile)
        
        GenerateComponent([lruname+'_Dist'],outdir)
# all the signal in one iom is impossible, for that afdx board may break.    
def GenerateconfigAFDX(mergedict,mode,devmap,split): # generate all the signal in one iom and cvt 
    templatedir = os.path.dirname(__file__)           
    #gac.cvts.clear()    
        
    lrus=[]
    if split:
        for msg in mergedict.inputafdxmsg.values():
            if msg['Lru'] not in lrus:
                lrus.append(msg['Lru'])
        for msg in mergedict.outputafdxmsg.values():
            if msg["Lru"] not in lrus:
                lrus.append(msg["Lru"])
    else:
        lrus.append(None)        
    print sorted(lrus)
    if mode == 'sim':
        
        inputmap = mergedict.inputafdxmsg
        outputmap = mergedict.outputafdxmsg
        inputvl = mergedict.inputvl
        outputvl = mergedict.outputvl
    else:
        inputmap = mergedict.outputafdxmsg
        outputmap = mergedict.inputafdxmsg
        inputvl= mergedict.outputvl
        outputvl= mergedict.inputvl      
    
    
    for lruname in lrus:
        #print lruname
        if lruname is not None:    
            cvts= CVTConnect(lruname+'_afdx')
            outfn = lruname + '_afdx'
        else:
            cvts= CVTConnect('SDIB'+'_afdx')
            outfn = 'SDIB' + '_afdx'
            
        if lruname is not None and lruname not in devmap.Map:
            print 'Lru: %s not exist in Device Map file'% lruname
            continue
        
        if devmap is not None:
            device = devmap.Map.get(lruname)[0]['SSDL_Device']
        else:
            device = 'AFDX-1'
                
        tmpl = Cheetah.Template.Template(file= os.path.join(templatedir, "ads2genAFDX.tpl"), 
                searchList=[{                    
                    "inputmodel" : inputmap,
                    'outputmodel': outputmap,
                    "LruName": lruname,                   
                    "device": device,
                    'inputvl': inputvl,
                    'outputvl':outputvl,
                    'mode': mode,
                    "cvts" : cvts,                    
                    }])
            
            
        fn = os.path.join(outdir, outfn + '.iom')
        temfile = tmpl.respond()
        open(fn, "w").write(temfile)
        
        tmpl = Cheetah.Template.Template(file= os.path.join(templatedir, "ads2genCVT.tpl"), 
                    searchList=[{                    
                        "cvtpoints" :cvts.cvts,
                        }])
    
        fn = os.path.join(outdir, outfn + '.cvt')
        temfile = tmpl.respond()
        open(fn, "w").write(temfile)
        
        GenerateComponent([outfn],outdir)
        
        
       

def GenerateconfigCAN(mergedict,mode,devmap,split): # generate all the signal in one iom and cvt   
    templatedir = os.path.dirname(__file__)           
    #gac.cvts.clear()
    
    lrus=[]
    if split:
        for msg in mergedict.inputCanMsg.values():
            if msg['Lru'] not in lrus:
                lrus.append(msg['Lru'])
        for msg in mergedict.outputCanMsg.values():
            if msg["Lru"] not in lrus:
                lrus.append(msg["Lru"])
    else:
        lrus.append(None)
            
    #print sorted(lrus)    
       
    if mode == 'sim':
        
        inputmap = mergedict.inputCanMsg
        outputmap = mergedict.outputCanMsg

    else:
        inputmap = mergedict.outputCanMsg
        outputmap = mergedict.inputCanMsg

    #print ('len of inputmsg: %d'%len(inputmap))
    #print ('len of outputmsg: %d'%len(outputmap))
    for lruname in lrus:
        if lruname is not None:
            cvts= CVTConnect(lruname+'_can')
            outfn = lruname + '_can'
        else:
            cvts= CVTConnect('SDIB'+'_can')
            outfn = 'SDIB' + '_can'
        if devmap is not None:
            devmap_= devmap.Map
        else:
            devmap_ = 'CAN-1-1'
        tmpl = Cheetah.Template.Template(file= os.path.join(templatedir, "ads2genCAN.tpl"), 
                searchList=[{                    
                    "inputmodel" : inputmap,
                    'outputmodel': outputmap,
                    "LruName": lruname,                   
                    "device": devmap_,
                    "mode": mode,
                    "cvts" : cvts,                    
                    }])
        
            
        fn = os.path.join(outdir, outfn + '.iom')
        temfile = tmpl.respond()
        open(fn, "w").write(temfile)
        
        tmpl = Cheetah.Template.Template(file= os.path.join(templatedir, "ads2genCVT.tpl"), 
                    searchList=[{                    
                        "cvtpoints" :cvts.cvts,
                        }])
    
        fn = os.path.join(outdir, outfn + '.cvt')
        temfile = tmpl.respond()
        open(fn, "w").write(temfile)
        
        GenerateComponent([outfn],outdir)
        
        
def GenerateconfigA429(mergedict,mode,devmap,split): # generate all the signal in one iom and cvt   
    templatedir = os.path.dirname(__file__)           
    #gac.cvts.clear()
    
    lrus=[]
    if split:
        for msg in mergedict.input429Chl.values():
            if msg['Lru'] not in lrus:
                lrus.append(msg['Lru'])
        for msg in mergedict.output429Chl.values():
            if msg["Lru"] not in lrus:
                lrus.append(msg["Lru"])
    else:
        lrus.append(None)        
    #print sorted(lrus)    
    
    if mode == 'sim':
        
        inputmap = mergedict.input429Chl
        outputmap = mergedict.output429Chl

    else:
        inputmap = mergedict.output429Chl
        outputmap = mergedict.input429Chl

    for lruname in lrus:
        if lruname is not None:
            cvts= CVTConnect(lruname+'_429')
            outfn = lruname + '_429'
        else:
            cvts= CVTConnect('SDIB'+'_429')
            outfn = 'SDIB' + '_429'
        if devmap is not None:
            devmap_= devmap.Map
        else:
            devmap_ = 'A429-1,TX-1'
        tmpl = Cheetah.Template.Template(file= os.path.join(templatedir, "ads2genA429.tpl"), 
                searchList=[{                    
                    "inputmodel" : inputmap,
                    'outputmodel': outputmap,
                    "LruName": lruname,                   
                    "device": devmap_,
                    "mode": mode,
                    "cvts" : cvts,                    
                    }])
           
        fn = os.path.join(outdir, outfn + '.iom')
        temfile = tmpl.respond()
        open(fn, "w").write(temfile)
        
        tmpl = Cheetah.Template.Template(file= os.path.join(templatedir, "ads2genCVT.tpl"), 
                    searchList=[{                    
                        "cvtpoints" :cvts.cvts,
                        }])
    
        fn = os.path.join(outdir, outfn + '.cvt')
        temfile = tmpl.respond()
        open(fn, "w").write(temfile)
            
        GenerateComponent([outfn],outdir)

# only generate 8 iom and 8 cvt, it not easy to find signal according lru name    
def GenerateconfigByAFDXBoard(inputmap,outdir,gac):  #generate the iom and cvt by afdx board number. 
    templatedir = os.path.dirname(__file__)
    
    gac.cvts.clear()       
    for boardname in LRU_TO_BOARD.keys():
        gac.cvts.clear()
        gac.hfname = boardname
        tmpl = Template(file= os.path.join(templatedir, "ads2genAFDXV5.tpl"), 
                searchList=[{                    
                    "model" : inputmap,
                    "LruName": LRU_TO_BOARD[boardname],                   
                    "device": boardname,
                    "gac" : gac,
                    }])
        fn = os.path.join(outdir, boardname + '.iom')
        temfile = tmpl.respond()
        open(fn, "w").write(temfile)
        
        tmpl = Template(file= os.path.join(templatedir, "ads2genCVT.tpl"), 
                searchList=[{                    
                    "cvtpoints" :gac.cvts,
                    }])

        fn = os.path.join(outdir, boardname + '.cvt')
        temfile = tmpl.respond()
        open(fn, "w").write(temfile)


def GenerateLRUExcel(inputmap,outdir,gac):    
        
    lrus = []
    for msg in gac.afdxmsg.values():
        lru = msg['LRU']
        if lru not in lrus:
            lrus.append(lru)   
    Lruexcel = Workbook()
    ws = Lruexcel.create_sheet(title='Lru-Name')
    # write header     
    lrus=sorted(lrus)
    for rowid in range(0,len(lrus)):
        ws.append([rowid,lrus[rowid]])
    
    # write data 
    
    Lruexcel.save('LruExcel-Name.xlsx')        


def GetAFDXBoardByLru(lru):
    for baord, lrulists in LRU_TO_BOARD.items():
        if lru in lrulists:
            return baord
    return None


def GenerateSession(hwcmps,simcmps,DPC,pnls,filename,outdir):#cmps is the list that included in the session, fileanme is the name of the session file
    templatedir = os.path.dirname(__file__)
    
      
    tmpl = Template(file= os.path.join(templatedir, "ads2genSess.tpl"), 
            searchList=[{                    
                "hwcmps" :hwcmps,
                "simcmps" : simcmps,
                "DPC" : DPC,
                "pnls": pnls
                }])

    fn = os.path.join(outdir, filename + '.ses')
    temfile = tmpl.respond()
    open(fn, "w").write(temfile)
        
def GenerateconfigbyforSSDL(inputmap,outdir,platform,gac):    
    templatedir = os.path.dirname(__file__)
    
    lrus = []
    for msg in gac.afdxmsg.values():
        lru = msg['LRU']
        if lru not in lrus:
            lrus.append(lru)   
    #print lrus
    lrus.sort()
    gac.cvts.clear()       
    for lru in lrus:
        afdxboard = GetAFDXBoardByLru(lru) 
        if afdxboard is None:
            logger.error('Lru %s do not blong to any afdx board, assign it to board FSIB_AFDX_08' % lru)
            afdxboard = "FSIB_AFDX_08"
        gac.cvts.clear()
        gac.hfname = lru
        #generate iom
        tmpl = Template(file= os.path.join(templatedir, "ads2genAFDXV4.tpl"), 
                searchList=[{                    
                    "model" : inputmap,
                    "LruName": lru,                   
                    "device": afdxboard,
                    "gac" : gac,
                    'hfname':None
                    }])
        fn = os.path.join(outdir, lru + '.iom')
        temfile = tmpl.respond()
        open(fn, "w").write(temfile)
        
        # generate CVT
        tmpl = Template(file= os.path.join(templatedir, "ads2genCVT.tpl"), 
                searchList=[{                    
                    "cvtpoints" :gac.cvts,
                    }])

        fn = os.path.join(outdir, lru + '.cvt')
        temfile = tmpl.respond()
        open(fn, "w").write(temfile)
        
        #generate cmp  one lru generate one cmp
        tem = []
        tem.append(lru)
        GenerateComponent(tem, lru,outdir)       
        
    #generate cmp by afdx board    
    for board, lrulist in LRU_TO_BOARD.items():
        GenerateComponent(lrulist, board,outdir)
        
    #generate ses config file
    hwcmps = [hwcmp+'.cmp' for hwcmp in lrus if hwcmp not in DS_Internal_Lru]
    sessionname = 'FSIB-Stimulation'
    simcmps = []
    pnls = [] 
    DPCname = None
    if platform == 'SSDL#1':
        DPCname = 'FSIB-DPC-01-03'
    elif platform == 'SSDL#2':
        DPCname = 'FSIB-DPC-01-09'
    else:
        logger.error('Platform name is wrong, it must be SSDL#1 or SSDL#2')
              
    GenerateSession(hwcmps,simcmps,DPCname,pnls,sessionname,outdir)




#used for cvt generate 
TypeIcdToads2 ={'SPAR': ('int32',4),
                 'DIS': ('int32',4),
                 'BOOL':('int32',4),
                 'BNR': ('real32',4),
                 'FLOAT':('real32',4),
                 'A429SDI':('int32',4),
                 'BLK':('byte',1),
                 'COD':('int32',4), 
                 'SINT':('int32',4), 
                 'INT':('int32',4),
                 'RESV':('int32',4), 
                 'A429_SSM_DIS':('int32',4), 
                 'UINT':('uint32',4), 
                 'CHAR':('int32',4), 
                 'UBNR':('real32',4), 
                 'OPAQUE':('byte',1), 
                 'BCD':('real32',4), 
                 'UBCD':('real32',4),
                 'ISO-5':('int32',4), 
                 'PAD':('int32',4), 
                 'A429_SSM_BNR':('int32',4), 
                 'A429PARITY':('int32',4), 
                 'A429_SSM_BCD':('int32',4), 
                 'A429OCTLBL':('int32',4),
                 'A429_SSM_CUSTOM':('int32',4),
                 'BYTES':  ('byte',      1),
                 'STRING': ('string',    1),}
                 

        
class CVTConnect(object):
    def __init__(self,filename):
        self.filename = filename
        self.cvts = dict()
        self.aliasname = dict()
    def newpoint(self,name,datatype,datasize,defaultvalue,elements=1,alias=[], varlength=0, channelmode="SAMPLING"):          
        return Bunch(name = name,datatype=datatype,datasize = datasize, defaultvalue = defaultvalue,elements=elements, alias=alias, varlength =varlength, channelmode=channelmode)
    
    def add_664_signal(self,msg,sig):
        message = msg
        name = None
        alias = []
        if sig['direct'] == 'Input':            
            finalpara=sig['Pubref'].split('.')[-1]       
            if finalpara.lower() in ['label','ssm','parity','pad','sdi']:
                lists=sig['Pubref'].split('.')
                res=ur"^L\d\d\d"
                for li in lists:
                    if re.search(res, li):
                        pos = lists.index(li)
                    #if len(sig['PubRefSrcName'].split('.')) > 3:
                        #print "pubref of sig is %s" % sig['Pubref']
                        wordname = sig['Pubref'].split('.')[pos]
                        name = message['fullname']+'.'+sig['DataSet']+'.'+wordname+'.'+sig['Signal']
                        #temp = sig['Pubref'].split('.')[0:pos+1]
                        #temp.insert(1,sig['txport'])
                        #name = '.'.join(temp) +'.'+finalpara
                        break
                    else:
                        #print "Pubref of signal is %s" % sig['Pubref']
                        wordname = sig['Pubref'].split('.')[-2]
                        name = message['fullname']+'.'+sig['DataSet']+'.'+wordname+'.'+sig['Signal']
                        #temp = sig['Pubref'].split('.')
                        #temp.insert(1,sig['txport'])
                        #name ='.'.join(temp)
                                
            else:
                name = message['fullname']+'.'+sig['DataSet']+'.'+sig['Signal']  # sig['Pubref'] #message['fullname']+'.'+sig['Signal']   # cvt point name can be all you like, but the alias must be pubref name,
            
            if msg['Lru'].startswith('RGW'):
                #print msg['Lru']+'.'+sig['Pubref']
                if (msg['Lru']+'.'+sig['Pubref']) not in Aliases:
                    alias.append(msg['Lru']+'.'+sig['Pubref'])
                    alias.append(msg['Lru']+'.'+sig['txport']+'.'+sig['Pubref'])
                    Aliases.append(msg['Lru']+'.'+sig['Pubref'])
                    self.aliasname[msg['Lru']+'.'+sig['Pubref']] = name
                else:
                    alias.append(msg['Lru']+'.'+sig['txport']+'.'+sig['Pubref'])
                    aliascvt = self.cvts[self.aliasname[msg['Lru']+'.'+sig['Pubref']]].alias
                    if msg['Lru']+'.'+sig['Pubref'] in aliascvt:
                        aliascvt.remove(msg['Lru']+'.'+sig['Pubref'])
                    ana = '#*'
                    for a in aliascvt:
                        if a.endswith('.'+sig['Pubref']) and a.startswith(msg['Lru']+'.'):
                            ana = a

                    print "@@@@@@@@%s not in alias library, you should use the name that added txport: %s or %s \n"% (msg['Lru']+'.'+sig['Pubref'],msg['Lru']+'.'+sig['txport']+'.'+sig['Pubref'],ana)

            else:
                alias.append(sig['Pubref'])
        elif sig['direct'] == 'Output':
            name = message['fullname']+'.'+sig['DataSet']+'.'+sig['Signal']
            alias.append(name)
        #else:
        #    alias.append(sig['PubRefSrcName'])
        #    alias.append(sig['pubportrefname'])
        
        #alias.append(sig['Pubref'])
        #alias.append(sig['pubportrefname'])
        
        #alias.append(name)
        #print sig['fullname']
        #print sig['SigType']
        #dtype, dsize = TypeIcdToads2[sig['SigType']]
        dtype = "byte"
        dsize = 1
        #msg['RxLength']
        defaultvalue = 0
        #if 'arinc 661' in message['MsgDataProtocolType'].lower():
            #chmode="MUXFIFO"
        #else:
            #chmode="SAMPLING"
        chmode="SAMPLING"
        newpoint = self.newpoint(name=name, datatype=dtype, datasize=dsize, defaultvalue=defaultvalue,alias = alias, channelmode=chmode)
        newpoint.elements = msg['RxLength']
        #if sig['SigType'] == "BYTES" or sig['SigType'] == "STRING":
            #newpoint.elements = sig['SigSize'] / 8
            #newpoint.varlength = 1
        self.cvts[name] = newpoint 
        return self.filename + '::' + name
    
    def add_664_fsb(self,msg,sig):
        message = msg
        name = message['fullname']+'.'+sig['DataSet']+'.'+'__fsb__'   # cvt point name can be all you like, but the alias must be pubref name,
        
        alias = []
        
        if sig['direct'] == 'Input':
            
            if msg['Lru'].startswith('RGW'):
                if (msg['Lru']+'.'+sig['Pubref']+'.'+'__fsb__') not in Aliases:
                    alias.append(msg['Lru']+'.'+sig['Pubref']+'.'+'__fsb__')
                    alias.append(msg['Lru']+'.'+sig['txport']+'.'+sig['Pubref']+'.'+'__fsb__')
                    Aliases.append(msg['Lru']+'.'+sig['Pubref']+'.'+'__fsb__')
                    self.aliasname[msg['Lru']+'.'+sig['Pubref']+'.'+'__fsb__'] = name
                else:
                    alias.append(msg['Lru']+'.'+sig['txport']+'.'+sig['Pubref']+'.'+'__fsb__')
                    
                    aliascvt = self.cvts[self.aliasname[msg['Lru']+'.'+sig['Pubref']+'.'+'__fsb__']].alias
                    if msg['Lru']+'.'+sig['Pubref']+'.'+'__fsb__' in aliascvt:
                        aliascvt.remove(msg['Lru']+'.'+sig['Pubref']+'.'+'__fsb__')
                    ana = '#*'
                    for a in aliascvt:
                        if a.endswith('.'+sig['Pubref']+'.'+'__fsb__') and a.startswith(msg['Lru']+'.'):
                            ana = a
                    
                    print "@@@@@@@@%s not in alias library, you should use the name that added txport: %s or %s \n"% (msg['Lru']+'.'+sig['Pubref']+'.'+'__fsb__',msg['Lru']+'.'+sig['txport']+'.'+sig['Pubref']+'.'+'__fsb__',ana)
                    
                
    
            else:
                alias.append(sig['Pubref']+'.'+'__fsb__')
                #alias.append(sig['pubportrefname']+'.'+'__fsb__')
        elif sig['direct'] == 'Output':
            alias.append(message['fullname']+'.'+sig['DataSet']+'.'+sig['Signal']+'.__fsb__')

        
        point = self.cvts.get(name)
        if point is None:            
            dtype = 'int8'
            dsize = 1
            defaultvalue = 0x00
            #if msg['Lru'].startswith('RGW'):
            #    alias.append(msg['Lru']+'.'+name)
            #else:
            #    alias.append(name)
            newpoint = self.newpoint(name=name, datatype=dtype, datasize=dsize, defaultvalue=defaultvalue,alias = alias)
     
            self.cvts[name] = newpoint 
            return self.filename + '::' + name
        else:
            point.alias.extend(alias)
            point.alias = list(set(point.alias)) # delete duplicate
            return None
    
    def add_664429_ssm(self,msg,sig):
        message = msg
        if sig['direct'] =='Input':
            name = message['fullname']+'.'+sig['Signal']+'.'+'SSM'  # cvt point name can be all you like, but the alias must be pubref name,
            #name = message['fullname']+'.'+'SSM' 
            #print ("come into 664429ssm , name is %s" % name)
            alias = []
        
        elif sig['direct'] == 'Output':
            pass
        '''
        if msg['LRU'].startswith('RGW'):
            alias.append(msg['LRU']+'.'+sig['PubRefSrcName']+'.'+'SSM')
            alias.append(msg['LRU'] + '.'+sig['pubportrefname']+'.'+'SSM')
        else:
            alias.append(sig['PubRefSrcName']+'.'+'SSM')
            alias.append(sig['pubportrefname']+'.'+'SSM')
        '''
        point = self.cvts.get(name)
        if point is None:            
            dtype = 'int32'
            dsize = 4
            defaultvalue = 0
            newpoint = self.newpoint(name=name, datatype=dtype, datasize=dsize, defaultvalue=defaultvalue,alias=alias)        

            self.cvts[name] = newpoint 
            return self.filename + '::' + name 
        else:
            point.alias.extend(alias)
            return None       
        
    '''
    def add_664_msg(self, msg):
        name = message['fullname']
        #alias = "apexport_a664_"+str(message.portId)
        alias = str(message.portName)
        if("out" in message.direction):
            alias = str(message.lruName) + "." + str(message.portName)
        if 'arinc 661' in $m['MsgDataProtocolType'].lower():
            chmode="MUXFIFO"
        else:
            chmode="SAMPLING"		
        attribs = self.newpoint(name=name, datatype="byte", elements=message.msgLength,  varlength=1, aliases=[alias], channelmode=chmode)
        self.cvtpoints[name] = attribs
        return self.cvtname + '::' + name
    '''    
    def add_664_control(self,msg,function):
        
        name = msg['fullname']+'.'+function
        
        dtype = 'int32'
        dsize = 4
        defaultvalue = 0x00
        if function=='__CRC_VAL__':
            defaultvalue = "0xFFFFFFFF"
        if function=='__CRC_FSB__':
            defaultvalue = "0x3"
        alias = [name]
        newpoint = self.newpoint(name=name, datatype=dtype, datasize=dsize, alias = alias,defaultvalue=defaultvalue)

        self.cvts[name] = newpoint 
        return self.filename + '::' + name  
        
        
    def add_monitor_signal(self,msg):
        
        name = msg['fullname']+'.'+ "Match"
        
        dtype = 'uint32'
        dsize = 4
        defaultvalue = msg['SourceUDP']
        alias = [name]
        newpoint = self.newpoint(name=name, datatype=dtype, datasize=dsize, alias = alias,defaultvalue=defaultvalue)

        self.cvts[name] = newpoint 
        return self.filename + '::' + name  
        
        
        #self.InputAfdxMessages = hfexecl.sheet_by_name('InputAfdxMessages')
        #InputSignals = hfexecl.sheet_by_name('InputSignals')
    def add_a825_msg_control(self,msg,function):
        
        name = msg['fullname']+'.'+function
        
        dtype = 'int32'
        dsize = 4
        defaultvalue = 0x00
        alias = [name]
        newpoint = self.newpoint(name=name, datatype=dtype, datasize=dsize, alias = alias,defaultvalue=defaultvalue)

        self.cvts[name] = newpoint 
        return self.filename + '::' + name  
    
    def add_a825_signal(self,msg,sig):
        message = msg
        name = None
        alias = []
        if sig['direct'] == 'Input':            
            
            name =message['fullname']+'.'+sig['RpName']#sig['Pubref'] #message['fullname']+'.'+sig['Signal']   # cvt point name can be all you like, but the alias must be pubref name,
            
            if msg['Lru'].startswith('RGW'):
                alias.append(msg['Lru']+'.'+sig['Pubref'])
                #alias.append(msg['Lru']+'.'+sig['pubportrefname'])
                #alias.append(msg['Lru']+'.'+name)
            else:
                alias.append(sig['Pubref'])
        elif sig['direct'] == 'Output':
            name = message['fullname']+'.'+sig['DpName']
            alias.append(name)
        #else:
        #    alias.append(sig['PubRefSrcName'])
        #    alias.append(sig['pubportrefname'])
        
        #alias.append(sig['Pubref'])
        #alias.append(sig['pubportrefname'])
        
        #alias.append(name)
        
        dtype, dsize = TypeIcdToads2[sig['SigType']]
        defaultvalue = 0
        newpoint = self.newpoint(name=name, datatype=dtype, datasize=dsize, defaultvalue=defaultvalue,alias = alias)
        if sig['SigType'] == "BYTES" or sig['SigType'] == "STRING":
            newpoint.elements = sig['SigSize'] / 8
            newpoint.varlength = 1
        self.cvts[name] = newpoint 
        return self.filename + '::' + name
    
    
    def add_a429_signal(self,msg,sig):
        message = msg
        name = None
        alias = []
        if sig['direct'] == 'Input':            
            
            if '_sdi_' in sig['Signal'].lower() or 'Source_Destination_ID' in sig['Signal']:
                endname ='SDI'
            elif '_ssm_' in sig['Signal'].lower() or 'Sign_Status_Matrix' in sig['Signal']:
                endname ='SSM'   
            else:
                endname = sig['Signal']
            
            
            name =message['fullname']+'.'+endname#sig['Pubref'] #message['fullname']+'.'+sig['Signal']   # cvt point name can be all you like, but the alias must be pubref name,
            
            if msg['Lru'].startswith('RGW'):
                alias.append(msg['Lru']+'.'+sig['Pubref'])
                #alias.append(msg['Lru']+'.'+sig['pubportrefname'])
                #alias.append(msg['Lru']+'.'+name)
            else:
                alias.append(sig['Pubref'])
        elif sig['direct'] == 'Output':
            name = message['fullname']+'.'+sig['Signal']
            alias.append(name)
        #else:
        #    alias.append(sig['PubRefSrcName'])
        #    alias.append(sig['pubportrefname'])
        
        #alias.append(sig['Pubref'])
        #alias.append(sig['pubportrefname'])
        
        #alias.append(name)
        dtype, dsize = TypeIcdToads2[sig['SigType']]
        defaultvalue = 0
        newpoint = self.newpoint(name=name, datatype=dtype, datasize=dsize, defaultvalue=defaultvalue,alias = alias)
        if sig['SigType'] == "BYTES" or sig['SigType'] == "STRING":
            newpoint.elements = sig['SigSize'] / 8
            newpoint.varlength = 1
        self.cvts[name] = newpoint 
        return self.filename + '::' + name

    def add_a429_msg_control(self,msg,function):
        
        name = msg['fullname']+'.'+function
        
        dtype = 'int32'
        dsize = 4
        defaultvalue = 0x00
        alias = [name]
        newpoint = self.newpoint(name=name, datatype=dtype, datasize=dsize, alias = alias,defaultvalue=defaultvalue)

        self.cvts[name] = newpoint 
        return self.filename + '::' + name 


    def add_Discrete(self,msg,sig):
        
        name = msg+'.'+sig
        alias = []

         
        dtype = 'int32'
        dsize = 4
        defaultvalue = 0
        alias.append(name)
        newpoint = self.newpoint(name=name, datatype=dtype, datasize=dsize, defaultvalue=defaultvalue,alias = alias)

        newpoint.elements = 0
        newpoint.varlength = 1
        self.cvts[name] = newpoint 
        return self.filename + '::' + name
    

"""
class GenerateAds2Conf(object):
    def __init__(self,workdir,outdir):    
        
        
        self.workdir = workdir
        self.outdir = outdir
        self.hfname = None
        self.afdxmsg = dict()
        self.afdxsignal = dict()
        self.vl = dict()
        self.hfexecl = None#xlrd.open_workbook(os.path.join(workdir,hfname)+'-icd.xlsx')
        self.cvts = dict()
        
    def sethfname(self,hfname):
        self.hfname = hfname
    def openfile(self,workdir,hfname):
        self.hfexecl= xlrd.open_workbook(os.path.join(workdir,hfname)+'-icd.xlsx')


"""

def DeteleMegSig(AnalysisRecord, hfnamelist):
    for key,vl in AnalysisRecord.inputafdxmsg.items():
        if vl['Lru'] in hfnamelist:
            del AnalysisRecord.inputafdxmsg[key]
    for key, vl in AnalysisRecord.inputCanMsg.items():
        if vl['Lru'] in hfnamelist:
            del AnalysisRecord.inputCanMsg[key]
    for key, vl in AnalysisRecord.inputafdxsignal.items():
        if vl['Lru'] in hfnamelist:
            del AnalysisRecord.inputafdxsignal[key]  
    for key,vl in AnalysisRecord.inputCanSignal.items():
        if vl['Lru'] in hfnamelist:
            del AnalysisRecord.inputCanSignal[key]       
    for key, vl in AnalysisRecord.inputvl.items():    
        if vl.lru in hfnamelist:
            del AnalysisRecord.inputvl[key]    
"""
class GenerateExcel(object):       
    def __init__(self):    
        self.inputafdxmsg = dict()
        self.inputafdxsignal = dict()
        self.outputafdxmsg = dict()
        self.outputafdxsignal = dict()
        self.inputCanMsg = dict()
        self.inputCanSignal = dict()
        self.outputCanMsg = dict()
        self.outputCanSignal = dict()  
        self.input429Signal = dict()
        self.input429Chl = dict()   
        self.output429Signal = dict()
        self.output429Chl = dict()    
        
    inSigAFDXColumns = (
          # header,                    value,                   len
            ("Skip",                   "skip",                        5),
            ("RP",                     "name",                       40),
            #("DataFormatType",         "dataformattype",             40),
            ("Pub_Ref",                "pubref",                     40),
            ("HostedFunction",         "LruName",                    20),
            ("Original Soruce LRU",    "originallru",                20),
            ("PortName",            "Portname",                    40),
            ("A664Message",            "MsgName",                    40),
            ("DS",                     "DsName",                     40),
            ("DP",                     "SigName",                    30),
            ("Validity",        "SourceSelection",            15),  #how to define this para
            ("ByteOffsetFSF",          "ByteOffsetFSB",              10),
            ("ByteOffsetWithinMsg",    "ByteOffsetDS",               10),
            ("DataSetSize",            "ByteSizeDS",                 10),
            ("DataFormatType",                "Encoding",                   10),
            ("BitOffsetWithinDS",      "BitOffset",                  10),
            ("ParameterSize",          "BitSize",                    10),
            ("LsbRes",                 "LsbValue",                   10),
            ("SysLatencyWCLimit",      "sysLatencywclimit",          10),
            ("Multiplier",             "multiplier",                 10),
            ("Label",                  "label",                      10),
            ("TransmissionIntervalMinimum",    "transmissionintervalminimum",           10),
            ("CodedSet",               "codedset",                   10),
            ("FullScaleRngMax",      "fullscalerangemax",          10),
            ("FullScaleRngMin",      "fullscalerangemin",          10),
            ("FuncRngMax",     "functionalrangemax",         10),
            ("FuncRngMin",     "functionalrangemin",         10),
            ("PublishedLatency",       "publishedlatency",           10),
            ("Units",                  "units",                      10),
            ("OneState",               "onestate",                   20),
            ("ZeroState",              "zerostate",                  20),
            ("RIUTemplate",           "RIUTemplate",                  10),
            ("UDC",                   "UDC",                          20),
            ("ErrorCode",             "ErrorCode",                    20),
            ("Comment",    "comment",           30),
      )                  
          
    
    outSigAFDXColumns = (
          # header,         value,                len
            ("Skip",        "skip",                5),
            ("DP",      "name",               40),
            ("HostedFunction",         "lru",                30),
            ("DS",     "DsName",             40),
            ("A664Message",     "msgname",            40),
            ("Validity",    "Validity",           15),
            ("ByteOffsetFSF",   "ByteOffsetFSB",      10),
            ("ByteOffsetWithinMsg",    "ByteOffsetDS",       10),
            ("DataSetSize",      "ByteSizeDS",         10),
            ("DataFormatType",     "Encoding",           10),
            ("BitOffsetWithinDS",   "BitOffset",          10),
            ("ParameterSize",     "BitSize",            10),
            ("LsbRes",    "LsbValue",           10),
            ("Multiplier",    "multiplier",           10),                 #need to add 
            ("Label",     "label",            10),
            ("TransmissionIntervalMinimum",    "transmissionintervalminimum",           15),
            ("CodedSet",   "codedset",      10),
            ("FullScaleRangeMax",    "fullscalerangemax",       10),
            ("FullScaleRangeMin",      "fullscalerangemin",         10),
            ("FunctionalRangeMax",     "functionalrangemax",           10),
            ("FunctionalRangeMin",   "functionalrangemin",          10),
            ("PublishedLatency",     "publishedlatency",            10),
            ("Units",    "units",           10),          
            ("OneState",   "onestate",          10),
            ("ZeroState",     "zerostate",            10),
            ("Consumer",    "consumer",           10),
            ("ErrorCode",             "ErrorCode",                    20),
            ("Comment",    "comment",           30),
      )    
    
    outSigCANColumns = (
          # header,         value,                len
            ("Skip",        "skip",                5),
            ("HostedFunction",         "lru",                30),
            ("CANMessage",     "msgname",            40),
            ("DP",      "name",               40),
            ("Validity",    "Validity",           15),
            ("DataFormatType",     "Encoding",           10),
            ("BitOffsetWithinDS",   "BitOffset",          10),
            ("ParameterSize",     "BitSize",            10),
            ("LsbRes",    "LsbValue",           10),  
            ("TransmissionIntervalMinimum",    "transmissionintervalminimum",           15),
            ("PublishedLatency",     "publishedlatency",            10),
            ("CodedSet",   "codedset",      10),
            ("FullScaleRangeMax",    "fullscalerangemax",       10),
            ("FullScaleRangeMin",      "fullscalerangemin",         10),
            ("FunctionalRangeMax",     "functionalrangemax",           10),
            ("FunctionalRangeMin",   "functionalrangemin",          10),
            ("Units",    "units",           10),          
            ("OneState",   "onestate",          10),
            ("ZeroState",     "zerostate",            10),
            ("ErrorCode",             "ErrorCode",                    20),
            ("Comment",    "comment",           30),        
      )
      
    inSigCANColumns = (
          # header,         value,                len
            ("Skip",                   "skip",                        5),
            ("RP",                     "name",                       40),
            ("Pub_Ref",                "pubref",                     40),
            ("HostedFunction",         "LruName",                    20),
            ("CANMessage",            "MsgName",                    40),
            ("MessageSize",            "msgsize",                    10),
            ("DP",                     "SigName",                    30),
            ("Validity",        "SourceSelection",            15),  #how to define this para     
            ("DataFormatType",                "Encoding",                   10),
            ("BitOffsetWithinDS",      "BitOffset",                  10),
            ("ParameterSize",          "BitSize",                    10),
            ("LsbRes",                 "LsbValue",                   10),
            ("SysLatencyWCLimit",      "sysLatencywclimit",          10),
            ("TransmissionIntervalMinimum",    "transmissionintervalminimum",           10),
            ("CodedSet",               "codedset",                   10),
            ("FullScaleRngMax",      "fullscalerangemax",          10),
            ("FullScaleRngMin",      "fullscalerangemin",          10),
            ("FuncRngMax",     "functionalrangemax",         10),
            ("FuncRngMin",     "functionalrangemin",         10),
            ("Units",                  "units",                      10),
            ("OneState",               "onestate",                   20),
            ("ZeroState",              "zerostate",                  20),
            ("ErrorCode",             "ErrorCode",                    20),
            ("Comment",    "comment",           30),
    
      )
     
         
    msgAfdxRxColumns = (
          # header,               value,           width
            ("Skip",              "skip",           5),
            ("ATA",               "atalist",              15),
            ("HostedFunction",               "lru",           30),
            ("Original Source LRU",   'lrulist',    20),
            ("A664Message",           "msgname",       40),
            #("MessageSize",            "txlength",        10),
            ("RxMessageSize",          "rxlength",      10),
            ("TxMessageSize",          "txlength",      10),
            ("MessageOverhead",          "overhead",      10),
            ("PortType",              "type",          10),
            ("QueueLength",       "queuelength",   10), 
            ("Pub_RefreshPeriod",              "txrate",          10),
            ("Sub_SamplePeriod",            "rxsamrate",        10),
            ("Sub_RefreshPeriod",            "rxrefrate",        10),
            ("ComPortRxID",            "portid",        10),
            ("Sub_Name",          "portname",      20), 
            ("VirtualLinkID",              "vlid",          10),
            ("SubVLID",             "subvlid",         10),
            ("BAG",               "BAG",           10),
            ("MTU",               "MTU",           10),
            ("EdeEnable",        "edeenable",    10),
            ("ComPortTxID",       "edesourceid",   10),
            ("Sub_IpAddress",            "destip",        15),
            ("Sub_UdpDstId",           "udpdstid",       15),
            ("Pub_MACAddress",         "sourecMACA",     15),
            #("SourceMACB",         "sourecMACB",     15),
            ("Pub_IpAddress",          "srcip",      15),
            ("Pub_UdpSrcId",         "udpsrcid",     10),          
            ("ActualPathA",        "ActualPathA",   60),
            ("ActualPathB",        "ActualPathB",   60),
            #("DestMAC",           "MACDest",      15)
            ("Networks",              "networks",           5),
            ("MsgDataProtocolType",               "msgdataprotocoltype",           30),
            ("TransmissionIntervalMinimum",           "transmissionintervalminimum",       10),
            ("ErrorCode",             "ErrorCode",                    20),
            ("Comment",    "comment",           30),
      )             
      
    msgAfdxTxColumns = (
          # header,               value,           width
            ("Skip",              "skip",           5),
            ("HostedFunction",               "lru",           30),
            ("A664Message",           "msgname",       40),   
            ("MessageSize",            "txlength",        10),
            ("MessageOverhead",          "overhead",      10),
            ("PortType",              "type",          10),
            ("QueueLength",       "queuelength",   10),
            ("ActivityTimeout",              "acttime",          10),
            ("Pub_RefreshPeriod",              "txrate",          10),
            ("ComPortRxID",            "portid",        10),
            ("Port_Name",          "portname",      20),
            ("VirtualLinkID",              "vlid",          10),
            ("SubVLID",             "subvlid",         10),
            ("BAG",               "BAG",           10),
            ("MTU",               "MTU",           10),
            ("EdeEnable",        "edeenable",    10),
            ("ComPortTxID",       "edesourceid",   10),
            ("Pub_MACAddress",         "sourecMACA",     15),
            #("SourceMACB",         "sourecMACB",     15),
            ("Pub_IpAddress",          "srcip",      15),
            ("Pub_UdpSrcId",         "udpsrcid",     10),
            ("Networks",              "networks",           5),
            ("MsgDataProtocolType",               "msgdataprotocoltype",           30),
            ("TransmissionIntervalMinimum",           "transmissionintervalminimum",       40),
            ("Sub_IpAddress",          "destip",      15),
            ("Sub_UdpDstId",         "udpdetiid",     10),
            ("ErrorCode",             "ErrorCode",                    20),
            ("Comment",    "comment",           30),
      )
    
    msgtxCanColumns = (
            ("Skip",              "skip",             5),
            ("HostedFunction",               "lru",             30),
            ("CANMessage",           "msgname",         40),
            ("MessageSize",            "txlength",        10),
            ("RefreshPeriod",              "txrate",          10),
            ("MessageID",          "txcanid",         10),
            ("Physical",          "txphysical",      20),
            ("ActivityTimeout",              "acttime",          10),
            ("CANMessageProtocolType",              "canprotype",          10),
            ("PartitionId",              "partitionid",          10),
            ("ErrorCode",             "ErrorCode",                    20),
            ("Comment",    "comment",           30),
      )
      
    msgrxCanColumns = (
            ("Skip",              "skip",             5),
            ("Pub_HostedFunction",               "lru",             30),
            ("Pub_CANMessage",           "msgname",         40),
            ("MessageSize",            "txlength",        10),
            ("Pub_RefreshPeriod",              "txrate",          10),
            ("Sub_SamplePeriod",              "rxrate",          10),
    
            #("TXCanMsgID",          "txcanid",         10),
            ("Pub_MessageID",          "rxcanid",         10),
            ("ActivityTimeout",              "txacttime",          10),
            ("CANMessageProtocolType",              "canprotype",          10),
            ("PartitionId",              "partitionid",          10),
            ("Pub_Physical",          "txphysical",      20),
            ("ErrorCode",             "ErrorCode",                    20),
            #("Physical",          "rxphysical",      20),
      )
      
    msgtx429Columns = (
            ("Skip",              "skip",             5),
            ("HostedFunction",               "lru",             30),
            ("A429Port",           "txportname",         40),
            ("A429Channel",           "txchlname",         40),          
            ("A429Message",           "msgname",         40),
            ("MessageSize",            "txlength",        10),
            ("RefreshPeriod",              "txrate",          10),
            #("MessageID/CanMsgID",          "txcanid",         10),
            ("Physical",          "txphysical",      20),
            #("ActivityTimeout",              "acttime",          10),
            ("MessageProtocolType",              "a429protype",          10),
            ("Label",              "Label",          10),
            ("SDI",               "SDI",            10),
            ("ErrorCode",             "ErrorCode",                    20),
            ("Comment",    "comment",           30),
      )
      
    msgrx429Columns = (
            ("Skip",              "skip",             5),
            ("Sub_A429Port",      'rxportname',      10),
            ("Sub_SamplePeriod",              "rxrate",          10),
            ("Sub_Physical",          "rxphysical",      20),
            ("Pub_HostedFunction",               "lru",             30),
            ("Pub_A429Port",           "txportname",         40),
            ("Pub_A429Channel",           "txchlname",         40),
            ("Pub_A429Word",           "msgname",         40),
            ("MessageSize",            "txlength",        10),
            ("Pub_RefreshPeriod",              "txrate",          10),
    
            ("A429ProtocolType",          "txprotocoltype",         10),
            ("Pub_Label",          "Label",         10),
            ('SDI',               "SDI",          10),          
            #("CANMessageProtocolType",              "canprotype",          10),
            #("PartitionId",              "partitionid",          10),
            ("Pub_Physical",          "txphysical",      20),
            ("ErrorCode",             "ErrorCode",                    20),
            #("Physical",          "rxphysical",      20),
      )
      
    outSig429Columns = (
          # header,         value,                len
            ("Skip",        "skip",                5),
            ("HostedFunction",         "lru",                30),
            ("A429Port",           "txportname",         40),
            ("A429Channel",           "txchlname",         40), 
            ("A429Message",     "msgname",            40),
            ("DP",      "name",               40),
            ("Validity",    "Validity",           15),
            ("DataFormatType",     "Encoding",           10),
            ("BitOffsetWithinDS",   "BitOffset",          10),
            ("ParameterSize",     "BitSize",            10),
            ("LsbRes",    "LsbValue",           10),  
            ("TransmissionIntervalMinimum",    "transmissionintervalminimum",           15),
            ("PublishedLatency",     "publishedlatency",            10),
            ("CodedSet",   "codedset",      10),
            ("FullScaleRangeMax",    "fullscalerangemax",       10),
            ("FullScaleRangeMin",      "fullscalerangemin",         10),
            ("FunctionalRangeMax",     "functionalrangemax",           10),
            ("FunctionalRangeMin",   "functionalrangemin",          10),
            ("Units",    "units",           10),          
            ("OneState",   "onestate",          10),
            ("ZeroState",     "zerostate",            10),
            ("ErrorCode",             "ErrorCode",                    20),
            ("Comment",    "comment",           30),        
      )
      
    inSig429Columns = (
          # header,         value,                len
            ("Skip",                   "skip",                        5),
            ("RP",                     "name",                       40),
            ("Pub_Ref",                "pubref",                     40),
            ("HostedFunction",         "LruName",                    20),
            ("Pub_A429Port",         "Portname",                    20),
            ("Pub_A429Channel",         "txchannel",                    20),
            ("A429Message",            "MsgName",                    40),
            #("MessageSize",            "msgsize",                    10),
            ("DP",                     "SigName",                    30),
            ("Validity",               "SourceSelection",            15),  #how to define this para     
            ("DataFormatType",         "Encoding",                   10),
            ("BitOffsetWithinDS",      "BitOffset",                  10),
            ("ParameterSize",          "BitSize",                    10),
            ("LsbRes",                 "LsbValue",                   10),
            ("SysLatencyWCLimit",      "sysLatencywclimit",          10),
            ("TransmissionIntervalMinimum",    "transmissionintervalminimum",           10),
            ("CodedSet",               "codedset",                   10),
            ("FullScaleRngMax",      "fullscalerangemax",          10),
            ("FullScaleRngMin",      "fullscalerangemin",          10),
            ("FuncRngMax",     "functionalrangemax",         10),
            ("FuncRngMin",     "functionalrangemin",         10),
            ("Units",                  "units",                      10),
            ("OneState",               "onestate",                   20),
            ("ZeroState",              "zerostate",                  20),
            ("ErrorCode",             "ErrorCode",                    20),
            ("Comment",    "comment",           30),
    
      )
        
    
    
    def formatOutput(self, outputfn):
    

        genExcelFile(outputfn, 
            (
                ("Input664Messages", self.msgAfdxRxColumns, 
                    sorted(self.inputafdxmsg, 
                           key=lambda msg: (msg.lru, msg.msgname)
                    )
                ),
    
                ("Input664Signals",  self.inSigAFDXColumns, 
                    sorted(self.inputafdxsignal,
                           key=lambda sig: (sig.LruName, sig.MsgName, sig.name)#sig.ByteOffsetDS, sig.BitOffset)
                    )
                ),
                ("Output664Messages", self.msgAfdxTxColumns, 
                    sorted(self.outputafdxmsg,
                           key=lambda msg: (msg.lru, msg.msgname)
                    )
                ),
                ("Output664Signals",  self.outSigAFDXColumns, 
                    sorted(self.outputafdxsignal,
                           key=lambda sig: (sig.lru, sig.msgname, sig.name)#sig.ByteOffsetDS, sig.BitOffset)
                    )
                ),
                ("Output825Messages", self.msgtxCanColumns, 
                    sorted(self.outputCanMsg),
                           key=lambda msg: (msg.lru, msg.msgname)
                    )
                ),
                ("Output825Signals",  self.outSigCANColumns, 
                    sorted(self.outputCanSignal,
                           key=lambda sig: (sig.lru, sig.msgname)
                    )
                ),
                ("Input825Messages", self.msgrxCanColumns, 
                    sorted(self.inputCanMsg,
                           key=lambda msg: (msg.lru, msg.msgname)
                    )
                ),
                ("Input825Signals",  self.inSigCANColumns, 
                    sorted(self.inputCanSignal,
                           key=lambda sig: (sig.LruName, sig.MsgName)
                    )
                ),               
                ("Output429Messages", self.msgtx429Columns, 
                    sorted(self.output429Chl,
                           key=lambda msg: (msg.lru, msg.msgname)
                    )
                ),
                ("Output429Signals",  self.outSig429Columns, 
                    sorted(self.output429Signal,
                           key=lambda sig: (sig.lru, sig.msgname)
                    )
                ),
                ("Input429Messages", self.msgrx429Columns, 
                    sorted(self.input429Chl,
                           key=lambda msg: (msg.lru, msg.msgname)
                    )
                ),
                ("Input429Signals",  self.inSig429Columns, 
                    sorted(self.input429Signal,
                           key=lambda sig: (sig.LruName, sig.MsgName)
                    )
                ),
            )
        )    

"""

def ProcessExcel(workdir,outdir, hfnamelist,platform,merge,mode,devmapfile,split):
    
    #first generate the config by HF independent.
    #gac = GenerateAds2Conf(workdir,outdir)
        
    if merge == True:
        #do the merge job and then generate ADS2 config
        if devmapfile is not None:
            devmap = ADS2_Device_Map(devmapfile)
            GenerateconfigDiscrete(devmap)
        else:
            devmap = None
        AnalysisRecord = AnalysisExcel.AnalysisExcel(workdir, hfnamelist)
        
        #generate one excel that list all the signals that contained in the IOM file 
        #geexcel = GenerateExcel()
        #geexcel.inputafdxmsg = AnalysisRecord.inputafdxmsg
        #geexcel.inputafdxsignal = AnalysisRecord.inputafdxsignal
        #geexcel.outputafdxmsg = AnalysisRecord.outputafdxmsg
        #geexcel.outputafdxsignal = AnalysisRecord.outputafdxsignal
        #geexcel.inputCanMsg = AnalysisRecord.inputCanMsg
        #geexcel.inputCanSignal = AnalysisRecord.inputCanSignal
        #geexcel.outputCanMsg = AnalysisRecord.outputCanMsg
        #geexcel.outputCanSignal = AnalysisRecord.outputCanSignal
        #geexcel.input429Signal = AnalysisRecord.input429Signal
        #geexcel.input429Chl = AnalysisRecord.input429Chl
        #geexcel.output429Signal = AnalysisRecord.output429Signal
        #geexcel.output429Chl = AnalysisRecord.output429Chl
        #geexcel.formatOutput(os.path.join(outdir,'SSDL1') + "-icd.xlsx")
        
        
        for vl in AnalysisRecord.inputvl.values():
            if int(vl.vlid) == 32775:
                print vl.receivelru
        #for vl in AnalysisRecord.outputvl.values():
        #    print vl.lru
        
        #Generate the Fanout switch config for 5.1, not including the RDIU1245
        Generate_Fanout_Config(outdir,devmap,AnalysisRecord.inputvl,AnalysisRecord.outputvl)
        #Generate the Fanout switch config for 5.2
        #Generate_Fanout_Config_52(outdir,devmap,AnalysisRecord.inputvl,AnalysisRecord.outputvl)  
        
        #delete the msg and signal that inter DS system
        DeteleMegSig(AnalysisRecord,hfnamelist)
        
        GenerateconfigAFDX(AnalysisRecord,mode,devmap,split)
        GenerateconfigCAN(AnalysisRecord,mode,devmap,split)
        GenerateconfigA429(AnalysisRecord,mode,devmap,split)
        #print "000000000000000000000000000000000000000000000000000000000000000000000"
    elif merge == False:
        #Generate ADS2 config for each hfname
        for hfname in hfnamelist:
            AnalysisRecord = AnalysisExcel.AnalysisExcel(workdir, [hfname])
        #mergedict = PerformMerge(AnalysisRecord)
        
            GenerateconfigAFDX(AnalysisRecord,mode,split)
            GenerateconfigCAN(AnalysisRecord,mode,split)
    else:
        print "merge must be true or false\n"
        return -1
'''    
def ProcessDictory(workdir,outdir, hfnamelist,platform,merge,mode):
    
    #first generate the config by HF independent.
    
    #gac = GenerateAds2Conf(workdir,outdir)
    
    hwcmps = []
    for hfname in hfnamelist:        
        gac.sethfname(hfname)
        #gac.openfile(workdir, hfname)
        hfdict = dsoutputdic.get(hfname)
        if hfdict is None:
            logger.error("in the ICD dict, does not have %s dict" % hfname)
            continue
        else:
            afdxmsgdict = hfdict.get("A664Msgs")
            gac.GetAfdxMessage(afdxmsgdict,False,hfname)
        
        hwcmps.append(hfname+'.cmp')
    
    #before generate the config, need to delete the duplicate DP sig in one msg, for that when we merge all the ICD into one excel, if the DP send to different and 
    #has different RP name, it will cuase the duplicate DP in msg
    
    ProcessDuplicateDPInMsg(gac.afdxmsg)
    
    #generate ses config file for config by hfname
    
             
    inputmap ={'afdx':gac.afdxmsg,               
           'vl':gac.vl,                              
           }
    GenerateLRUExcel(inputmap,outdir,)   #write the lru name into one excel file      
    GenerateconfigbyforSSDL(inputmap,os.path.join(outdir,platform,'By_LRU_Name'),platform,gac)  #generate config for ssdl by lru name, this is we use now
'''

def usage():
    print ('Input args is invalid.....')
    print ('-c -- print out the log message to the sterr')
    print ('-p -- print out the progress')
    print ('--wordir -- define the root directory of ICD need to import ana analysis ')
    print ('--outdir -- define the ICD analysis results output directory')
    print ('--haname -- define the name list that need to generate')
    print ('--loglevel -- set the loglevel=[TRACE|INFO|WARN|ERROR]')
    print ('--platform -- set the SSDL#1 or SSDL#2')
    print ('--inputtype -- set the inputfile, Excel files or Dictory generated by xiaochun tool')
    print ('you must do as what as up')

if __name__ == "__main__":
    opts, args = getopt.getopt(sys.argv[1:], 'cpms', ['inputtype=','workdir=','outdir=','hfname=','loglevel=','platform=','merge=','mode=','devmap='])
    
    loglevels={
               'TRACE':logger.TRACE,
               'INFO' :logger.INFO,
               'WARN' :logger.WARN,
               'ERROR':logger.ERROR}
    
    workdir = None
    outdir = None
    hfnamelist = None
    loglevel = loglevels.get('ERROR')
    logconsole = None
    logprogress = None
    platform = None
    inputtype = None
    merge = False
    mode = None
    devmapfile = None
    split = False
    
    if len(opts) == 0 :
        print ('the input opt can no be None')
        sys.exit(0)
    
    for o, v in opts:
        if o in ['-c','--console']:
            logconsole = True
            logprogress = False
        elif o in ['-p','--progress']:
            logconsole = False
            logprogress = True
        elif o in ['--workdir']:
            workdir = v
        elif o in ['--outdir']:
            outdir = v
        elif o in ['--inputtype']:            
            inputtype = v
        elif o in ['--hfname']:
            hfnamelist = v.split(',') 
        elif o in ['--loglevel']:
            loglevel = loglevels.get(v,logger.INFO) 
        elif o in ['--platform']:
            platform = v
        elif o in ['-m','--merge']:
            merge = True
        elif o in ['-s','--split']:
            split = True
        elif o in ['--mode']:
            mode = v
        elif o in ['--devmap']:
            devmapfile = v
        else:
            usage()
            sys.exit(0)
    
    
    logger.setup(level = loglevel, filename=os.path.join(outdir,sys.argv[0].split('\\')[-1][:-3]) + "-icd.log", console = logconsole, progress=logprogress)
    logger.info('Start the Ads2 configuration files generate....')
    
    if inputtype == 'Excel':
        ProcessExcel(workdir,outdir, hfnamelist,platform,merge,mode,devmapfile,split)
    elif inputtype == "Dictory":
        pass
        #ProcessDictory(workdir,outdir, hfnamelist,platform,merge,mode)
    else:
        print ("The inputtype must be Excel or Dictory...\n")
        sys.exit(0)   

    print("ADS2 CFG Generate OK")
    #GenerateSwitchConfig(dsoutputdic,os.path.join(outdir,platform,'Switch-Config'))
    '''
    hfhaname = dsoutputdic.keys()
    print hfhaname
    
    onehf = dsoutputdic.get('FDAS_L1')
    lrumsg = onehf.keys()    
    onemsg = onehf.get(lrumsg[0])
    msgkeys = onemsg.keys()
    print '\n'
    print msgkeys
    RPdic = onemsg.get('RP')
    rpsigkeys = RPdic.keys()
    if RPdic is not None:
        rp = RPdic.values()[0]
        rpkeys=rp.keys()
        print rpkeys
    '''
    
    #GenerateByHF(workdir,outdir,hfname, dsoutputdic,platform)
    #print ('Generate for HF finish...')
    #GenerateByLRU(workdir,outdir,hfname, dsoutputdic1,platform)  # mode=True: together mode=False: LRU  hfname=True: generate by hfname and together
    #print ('generate for LRU finish...')    
    #msg keys
    #['Network', 'IPDestAddrFormat', 'BAG', 'SourceUDP', 'EdeEnable', 'RP', 'RxLength', 'RxRate', 'TxPortName', 'OverHead', 'SourceIP', 'TxPortID', 'QueueLength', 'MTU', 'HardWare', 'DestIp', 'Length', 'SubVLID', 'VLID', 'Rate', 'LRU', 'PortName', 'Message']
    #['Network', 'IPDestAddrFormat', 'DestUDP', 'BAG', 'SourceUDP', 'EdeEnable', 'RP', 'RxLength', 'RxRate', 'TxPortName', 'RxPortID', 'OverHead', 'SourceIP', 'SourceMAC', 'TxPortID', 'QueueLength', 'MTU', 'HardWare', 'DestIp', 'Length', 'RxPortType', 'SubVLID', 'VLID', 'Rate', 'LRU', 'PortName', 'Message']
    #['PubRefSrcGuid', 'PubRefSrcName', 'RPGuid', 'FsbOffset', 'Signal', 'SignalSize', 'SignalType', 'DataSet', 'LRU', 'RPName', 'LsbValue', 'SignalOffset', 'RPNameDef', 'Message', 'DSSize', 'DSOffset', 'RPGuidDef']
    #rp KeysV
    #['PubRefSrcGuid', 'PubRefSrcName', 'RPGuid', 'FsbOffset', 'Signal', 'SignalSize', 'SignalType', 'DataSet', 'LRU', 'RPName', 'LsbValue', 'SignalOffset', 'RPNameDef', 'Message', 'DSSize', 'DSOffset', 'RPGuidDef']

#-c -m --loglevel=INFO  --outdir=D:\C919Tools\GeneExcelICD\ads2config\SSDLconfig --devmap=D:\C919Tools\ICDImportV2\DeviceMap.xlsx --mode=stim--inputtype=Excel --hfname=FDAS_L1,FDAS_L3,FDAS_R3,IMA_DM_L4,IMA_DM_L5,IMA_DM_R4,SYNOPTICMENUAPP_L,SYNOPTICMENUAPP_R,SYNOPTICPAGEAPP_L,SYNOPTICPAGEAPP_R,HF_IDULEFTINBOARD,HF_IDULEFTOUTBOARD,HF_IDUCENTER,HF_IDURIGHTINBOARD,HF_IDURIGHTOUTBOARD,HF_CCD1,HF_CCD2,HF_DCP1,HF_DCP2,HF_EVS,HF_HCU1,HF_HCU2,HF_HPU1,HF_HPU2,HF_ISIS,HF_MCMW1,HF_MCMW2,HF_MKB1,HF_MKB2,HF_RCP,HF_RLS1,HF_RLS2,ECL_L,ECL_R,VIRTUALCONTROLAPP_L,VIRTUALCONTROLAPP_R --workdir="D:\C919Tools\GeneExcelICD\ads2config"
#HF_RIU_1,HF_TCP_3,HF_EMPC_EPS,HF_CARGO_FIRECNTRLPANEL,HF_ENGINEAPU_FIRECNTRLPANEL,HF_FUELOVERHEADPANEL,HF_L_ID,HF_R_ID,HF_L_NAISWITCH,HF_R_NAISWITCH,HF_WHCA,HF_GAGEASSY,HF_MCMW1,HF_MCMW2,HF_LGCU2,HF_AIR_COND_CPA,HF_DIM_CTRL_PWR,HF_ICE_CABIN_LT_CPA,HF_INSTR_CPA_L,HF_INSTR_CPA_R,HF_EMERGENCYLIGHTINGSW,HF_ACU,HF_FUELCONTROLSW_R